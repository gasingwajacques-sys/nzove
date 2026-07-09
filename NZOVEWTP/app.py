import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta, date
import numpy as np
import os
import hashlib
import shutil
import json
from io import BytesIO
import random
import sqlite3
from contextlib import contextmanager

# ================= PAGE CONFIG =================
st.set_page_config(
    page_title="Nzove Water Treatment Plant - Monitoring System",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ================= DATABASE FUNCTIONS =================
DB_FILE = "nzove_database.db"

@contextmanager
def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

def init_database():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Users table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                username TEXT PRIMARY KEY,
                password TEXT NOT NULL,
                role TEXT NOT NULL,
                branch TEXT DEFAULT 'NZOVE',
                security_question TEXT,
                security_answer TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Records table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                branch TEXT NOT NULL,
                shift TEXT NOT NULL,
                date_range TEXT,
                team TEXT,
                inflow_m3 REAL DEFAULT 0,
                outflow_m3 REAL DEFAULT 0,
                backwashing_m3 REAL DEFAULT 0,
                non_revenue_pct REAL DEFAULT 10,
                revenue_water_m3 REAL DEFAULT 0,
                loss_m3 REAL DEFAULT 0,
                total_production_cost REAL DEFAULT 0,
                water_revenue REAL DEFAULT 0,
                net_profit REAL DEFAULT 0,
                recorded_by TEXT,
                efficiency REAL DEFAULT 0,
                ph_level REAL DEFAULT 7.0,
                turbidity_ntu REAL DEFAULT 0,
                temperature REAL DEFAULT 20.0,
                chlorine_mgL REAL DEFAULT 0.3,
                nitrate_mgL REAL DEFAULT 10.0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Team performance table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS team_performance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                branch TEXT NOT NULL,
                team_type TEXT NOT NULL,
                team_leader TEXT,
                shift TEXT NOT NULL,
                raw_water REAL DEFAULT 0,
                treated_water REAL DEFAULT 0,
                backwashing REAL DEFAULT 0,
                non_revenue_pct REAL DEFAULT 10,
                revenue_water REAL DEFAULT 0,
                expense REAL DEFAULT 0,
                revenue REAL DEFAULT 0,
                efficiency REAL DEFAULT 0,
                quality_score REAL DEFAULT 100,
                recorded_by TEXT,
                team_comment TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Meter readings table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS meter_readings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reading_date TEXT NOT NULL,
                branch TEXT NOT NULL,
                meter_type TEXT NOT NULL,
                meter_name TEXT NOT NULL,
                reading_value REAL DEFAULT 0,
                recorded_by TEXT,
                recorded_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Water quality table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS water_quality (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                datetime TEXT NOT NULL,
                branch TEXT NOT NULL,
                flow_m3 REAL DEFAULT 0,
                raw_ntu REAL DEFAULT 0,
                stock_pct REAL DEFAULT 0,
                dose_mgL REAL DEFAULT 0,
                outlet_ntu REAL DEFAULT 0,
                efficiency REAL DEFAULT 0,
                pump_Lh REAL DEFAULT 0,
                pump_Lmin REAL DEFAULT 0,
                raw_ph REAL DEFAULT 7.0,
                raw_temperature REAL DEFAULT 20.0,
                raw_chlorine REAL DEFAULT 0,
                raw_nitrate REAL DEFAULT 0,
                recorded_by TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Team info table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS team_info (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch TEXT NOT NULL,
                team_type TEXT NOT NULL,
                team_leader TEXT,
                members TEXT,
                representative TEXT,
                contact TEXT,
                updated_by TEXT,
                updated_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(branch, team_type)
            )
        ''')
        
        # Thresholds table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS thresholds (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch TEXT NOT NULL,
                threshold_key TEXT NOT NULL,
                threshold_value REAL DEFAULT 0,
                UNIQUE(branch, threshold_key)
            )
        ''')
        
        # Meters config table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS meters_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch TEXT NOT NULL,
                meter_type TEXT NOT NULL,
                meters TEXT,
                UNIQUE(branch, meter_type)
            )
        ''')
        
        # Monthly evaluation table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS monthly_evaluation (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                branch TEXT NOT NULL,
                team_type TEXT NOT NULL,
                team_leader TEXT,
                avg_efficiency REAL DEFAULT 0,
                total_raw REAL DEFAULT 0,
                total_output REAL DEFAULT 0,
                revenue_water REAL DEFAULT 0,
                backwashing REAL DEFAULT 0,
                non_revenue_pct REAL DEFAULT 0,
                total_cost REAL DEFAULT 0,
                total_revenue REAL DEFAULT 0,
                total_profit REAL DEFAULT 0,
                cost_per_m3 REAL DEFAULT 0,
                compliance_score REAL DEFAULT 0,
                rank INTEGER DEFAULT 0,
                remarks TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(year, month, branch, team_type)
            )
        ''')
        
        conn.commit()

# Initialize database
init_database()

# ================= SECURITY FUNCTIONS =================
def make_hashes(password):
    return hashlib.sha256(str.encode(password)).hexdigest()

def check_hashes(password, hashed_text):
    return make_hashes(password) == hashed_text

# ================= USER MANAGEMENT =================
def load_users():
    users = {}
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT username, password, role, branch FROM users")
        for row in cursor.fetchall():
            users[row[0]] = {
                "password": row[1],
                "role": row[2],
                "branch": row[3] if row[3] else "NZOVE"
            }
    
    if not users:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT OR REPLACE INTO users (username, password, role, branch) VALUES (?, ?, ?, ?)",
                ("admin", make_hashes("admin123"), "Admin", "NZOVE")
            )
            conn.commit()
        users["admin"] = {
            "password": make_hashes("admin123"),
            "role": "Admin",
            "branch": "NZOVE"
        }
    return users

def save_user(username, password, role, branch):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT OR REPLACE INTO users (username, password, role, branch) VALUES (?, ?, ?, ?)",
            (username, make_hashes(password), role, branch if branch else "NZOVE")
        )
        conn.commit()
        return True

def update_user_password(username, new_password):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET password = ? WHERE username = ?",
            (make_hashes(new_password), username)
        )
        conn.commit()
        return cursor.rowcount > 0

def delete_user(username):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE username = ? AND username != 'admin'", (username,))
        conn.commit()
        return cursor.rowcount > 0

def get_all_users():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT username, role, branch FROM users")
        return [{"username": row[0], "role": row[1], "branch": row[2] if row[2] else "NZOVE"} 
                for row in cursor.fetchall()]

def set_security_question(username, question, answer):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET security_question = ?, security_answer = ? WHERE username = ?",
            (question, make_hashes(answer.strip().lower()), username)
        )
        conn.commit()

def get_security_question(username):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT security_question, security_answer FROM users WHERE username = ?", (username,))
        row = cursor.fetchone()
        if row and row[0]:
            return row[0]
    return None

def verify_security_answer(username, answer):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT security_answer FROM users WHERE username = ?", (username,))
        row = cursor.fetchone()
        if row and row[0]:
            return check_hashes(answer.strip().lower(), row[0])
    return False

def reset_password_via_security(username, new_password):
    return update_user_password(username, new_password)

def change_user_password(username, old_password, new_password):
    users = load_users()
    if username in users and check_hashes(old_password, users[username]["password"]):
        return update_user_password(username, new_password)
    return False

# ================= TEAM FUNCTIONS =================
TEAM_TYPES = ["A", "B", "C", "D"]

def save_team_info(branch, team_type, team_leader, members, representative, contact, updated_by):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO team_info 
            (branch, team_type, team_leader, members, representative, contact, updated_by, updated_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (branch, team_type, team_leader, members, representative, contact, updated_by))
        conn.commit()

def get_team_info(branch, team_type):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM team_info WHERE branch = ? AND team_type = ?",
            (branch, team_type)
        )
        row = cursor.fetchone()
        if row:
            return dict(row)
    return None

def save_team_performance(date, branch, team_type, team_leader, shift, raw_water, treated_water, 
                          backwashing, non_revenue_pct, revenue_water, expense, revenue, 
                          efficiency, quality_score, recorded_by, team_comment=""):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO team_performance 
            (date, branch, team_type, team_leader, shift, raw_water, treated_water, 
             backwashing, non_revenue_pct, revenue_water, expense, revenue, 
             efficiency, quality_score, recorded_by, team_comment)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (str(date), branch, team_type, team_leader, shift, raw_water, treated_water, 
              backwashing, non_revenue_pct, revenue_water, expense, revenue, 
              efficiency, quality_score, recorded_by, team_comment))
        conn.commit()

def get_team_performance(branch, team_type, start_date=None, end_date=None):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        query = """
            SELECT * FROM team_performance 
            WHERE branch = ? AND team_type = ?
        """
        params = [branch, team_type]
        
        if start_date:
            query += " AND date >= ?"
            params.append(str(start_date))
        if end_date:
            query += " AND date <= ?"
            params.append(str(end_date))
        
        query += " ORDER BY date ASC"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        return pd.DataFrame([dict(row) for row in rows])

def delete_team_performance_for_date(date, branch):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "DELETE FROM team_performance WHERE date = ? AND branch = ?",
            (str(date), branch)
        )
        conn.commit()

# ================= METER FUNCTIONS =================
def save_meter_reading(reading_date, branch, meter_type, meter_name, reading_value, recorded_by):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO meter_readings 
            (reading_date, branch, meter_type, meter_name, reading_value, recorded_by)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (reading_date, branch, meter_type, meter_name, reading_value, recorded_by))
        conn.commit()

def get_latest_meter_reading(branch, meter_type, meter_name):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM meter_readings 
            WHERE branch = ? AND meter_type = ? AND meter_name = ?
            ORDER BY recorded_time DESC, reading_date DESC
            LIMIT 1
        ''', (branch, meter_type, meter_name))
        row = cursor.fetchone()
        if row:
            return dict(row)
    return None

def get_meter_readings_history(branch, meter_type=None, meter_name=None):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        query = "SELECT * FROM meter_readings WHERE branch = ?"
        params = [branch]
        
        if meter_type:
            query += " AND meter_type = ?"
            params.append(meter_type)
        if meter_name:
            query += " AND meter_name = ?"
            params.append(meter_name)
        
        query += " ORDER BY reading_date ASC, recorded_time ASC"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        return pd.DataFrame([dict(row) for row in rows])

# ================= WATER QUALITY FUNCTIONS =================
def save_water_quality_record(datetime_str, branch, flow_m3, raw_ntu, stock_pct,
                              dose_mgL, outlet_ntu, efficiency, pump_Lh, pump_Lmin,
                              raw_ph, raw_temperature, raw_chlorine, raw_nitrate, recorded_by):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO water_quality 
            (datetime, branch, flow_m3, raw_ntu, stock_pct, dose_mgL, outlet_ntu, 
             efficiency, pump_Lh, pump_Lmin, raw_ph, raw_temperature, raw_chlorine, 
             raw_nitrate, recorded_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (datetime_str, branch, flow_m3, raw_ntu, stock_pct, dose_mgL, outlet_ntu,
              efficiency, pump_Lh, pump_Lmin, raw_ph, raw_temperature, raw_chlorine,
              raw_nitrate, recorded_by))
        conn.commit()

def get_water_quality_records(branch, limit=100):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM water_quality 
            WHERE branch = ?
            ORDER BY datetime DESC
            LIMIT ?
        ''', (branch, limit))
        rows = cursor.fetchall()
        df = pd.DataFrame([dict(row) for row in rows])
        if not df.empty:
            df["datetime"] = pd.to_datetime(df["datetime"], errors='coerce')
        return df

# ================= RECORD FUNCTIONS =================
def save_record(record_data):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO records 
            (date, branch, shift, date_range, team, inflow_m3, outflow_m3, 
             backwashing_m3, non_revenue_pct, revenue_water_m3, loss_m3,
             total_production_cost, water_revenue, net_profit, recorded_by,
             efficiency, ph_level, turbidity_ntu, temperature, chlorine_mgL, nitrate_mgL)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            record_data.get("date", ""),
            record_data.get("branch", "NZOVE"),
            record_data.get("shift", ""),
            record_data.get("date_range", ""),
            record_data.get("team", ""),
            record_data.get("inflow_m3", 0),
            record_data.get("outflow_m3", 0),
            record_data.get("backwashing_m3", 0),
            record_data.get("non_revenue_pct", 10),
            record_data.get("revenue_water_m3", 0),
            record_data.get("loss_m3", 0),
            record_data.get("total_production_cost", 0),
            record_data.get("water_revenue", 0),
            record_data.get("net_profit", 0),
            record_data.get("recorded_by", ""),
            record_data.get("efficiency", 0),
            record_data.get("ph_level", 7.0),
            record_data.get("turbidity_ntu", 0),
            record_data.get("temperature", 20.0),
            record_data.get("chlorine_mgL", 0.3),
            record_data.get("nitrate_mgL", 10.0)
        ))
        conn.commit()

def get_records(branch, year=None, month=None, day=None):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        query = "SELECT * FROM records WHERE branch = ?"
        params = [branch]
        
        if year:
            query += " AND strftime('%Y', date) = ?"
            params.append(str(year))
        if month:
            query += " AND strftime('%m', date) = ?"
            params.append(f"{month:02d}")
        if day:
            query += " AND strftime('%d', date) = ?"
            params.append(f"{day:02d}")
        
        query += " ORDER BY date DESC"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        df = pd.DataFrame([dict(row) for row in rows])
        if not df.empty and "date" in df.columns:
            df["Date"] = pd.to_datetime(df["date"])
            if "shift" in df.columns:
                df["Shift"] = df["shift"]
        return df

def delete_record_for_date(date, branch, shift="ALL"):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        if shift == "ALL":
            cursor.execute(
                "DELETE FROM records WHERE date = ? AND branch = ?",
                (str(date), branch)
            )
        else:
            cursor.execute(
                "DELETE FROM records WHERE date = ? AND branch = ? AND shift = ?",
                (str(date), branch, shift)
            )
        deleted = cursor.rowcount
        conn.commit()
        if deleted > 0:
            delete_team_performance_for_date(date, branch)
        return deleted > 0

def delete_all_records():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM records")
        cursor.execute("DELETE FROM team_performance")
        cursor.execute("DELETE FROM water_quality")
        cursor.execute("DELETE FROM meter_readings")
        cursor.execute("DELETE FROM monthly_evaluation")
        conn.commit()

# ================= THRESHOLD FUNCTIONS =================
def load_thresholds_for_branch(branch_name):
    defaults = {
        "pH_min": 6.5, "pH_max": 8.5,
        "Turbidity_max": 5.0,
        "Temperature_min": 15.0, "Temperature_max": 25.0,
        "Chlorine_min": 0.2, "Chlorine_max": 0.5,
        "Nitrate_max": 50.0,
        "Efficiency_min": 75.0,
        "Price_Lime": 450.0,
        "Price_FL4440": 2500.0,
        "Price_NaCl": 300.0,
        "Price_Electricity": 148.0,
        "Non_Revenue_Pct": 10.0,
        "Base_Cost": 768.0
    }
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT threshold_key, threshold_value FROM thresholds WHERE branch = ?",
            (branch_name,)
        )
        rows = cursor.fetchall()
        current_map = {}
        for row in rows:
            current_map[row[0]] = row[1]
        
        for k, v in defaults.items():
            if k not in current_map:
                current_map[k] = v
        
        return current_map

def save_thresholds_for_branch(branch_name, threshold_dict):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        for k, v in threshold_dict.items():
            cursor.execute('''
                INSERT OR REPLACE INTO thresholds (branch, threshold_key, threshold_value)
                VALUES (?, ?, ?)
            ''', (branch_name, k, float(v)))
        conn.commit()

def load_global_base_cost():
    thresholds = load_thresholds_for_branch("NZOVE")
    return float(thresholds.get("Base_Cost", 768))

def save_global_base_cost(new_cost):
    thresholds = load_thresholds_for_branch("NZOVE")
    thresholds["Base_Cost"] = float(new_cost)
    save_thresholds_for_branch("NZOVE", thresholds)

# ================= METERS CONFIG =================
def load_meters_config(branch_name):
    defaults = {
        "raw_water": ["Meter 1", "Meter 2", "Meter 3"],
        "treated_water": ["Meter 1", "Meter 2", "Meter 3", "Meter 4", "Meter 5"],
        "backwashing": ["Backwash Meter 1", "Backwash Meter 2"],
        "electricity": ["Meter 1", "Meter 2", "Meter 3", "Meter 4", "Meter 5"]
    }
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT meter_type, meters FROM meters_config WHERE branch = ?",
            (branch_name,)
        )
        rows = cursor.fetchall()
        if rows:
            config = {}
            for row in rows:
                config[row[0]] = json.loads(row[1])
            return config
    
    # Save defaults
    with get_db_connection() as conn:
        cursor = conn.cursor()
        for meter_type, meters in defaults.items():
            cursor.execute('''
                INSERT OR REPLACE INTO meters_config (branch, meter_type, meters)
                VALUES (?, ?, ?)
            ''', (branch_name, meter_type, json.dumps(meters)))
        conn.commit()
    
    return defaults

def save_meters_config(branch_name, config_dict):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        for meter_type, meters in config_dict.items():
            cursor.execute('''
                INSERT OR REPLACE INTO meters_config (branch, meter_type, meters)
                VALUES (?, ?, ?)
            ''', (branch_name, meter_type, json.dumps(meters)))
        conn.commit()

def add_meter(branch_name, meter_type, meter_name):
    config = load_meters_config(branch_name)
    if meter_type in config:
        if meter_name not in config[meter_type]:
            config[meter_type].append(meter_name)
            save_meters_config(branch_name, config)
            return True
    return False

def remove_meter(branch_name, meter_type, meter_name):
    config = load_meters_config(branch_name)
    if meter_type in config:
        if meter_name in config[meter_type]:
            config[meter_type].remove(meter_name)
            save_meters_config(branch_name, config)
            return True
    return False

def edit_meter(branch_name, meter_type, old_name, new_name):
    config = load_meters_config(branch_name)
    if meter_type in config:
        if old_name in config[meter_type]:
            index = config[meter_type].index(old_name)
            config[meter_type][index] = new_name
            save_meters_config(branch_name, config)
            return True
    return False

def get_all_meters(branch_name):
    return load_meters_config(branch_name)

# ================= MONTHLY EVALUATION =================
def save_monthly_evaluation(evaluation_data):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO monthly_evaluation 
            (year, month, branch, team_type, team_leader, avg_efficiency, total_raw,
             total_output, revenue_water, backwashing, non_revenue_pct, total_cost,
             total_revenue, total_profit, cost_per_m3, compliance_score, rank, remarks)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            evaluation_data["year"],
            evaluation_data["month"],
            evaluation_data["branch"],
            evaluation_data["team_type"],
            evaluation_data.get("team_leader", ""),
            evaluation_data.get("avg_efficiency", 0),
            evaluation_data.get("total_raw", 0),
            evaluation_data.get("total_output", 0),
            evaluation_data.get("revenue_water", 0),
            evaluation_data.get("backwashing", 0),
            evaluation_data.get("non_revenue_pct", 0),
            evaluation_data.get("total_cost", 0),
            evaluation_data.get("total_revenue", 0),
            evaluation_data.get("total_profit", 0),
            evaluation_data.get("cost_per_m3", 0),
            evaluation_data.get("compliance_score", 0),
            evaluation_data.get("rank", 0),
            evaluation_data.get("remarks", "")
        ))
        conn.commit()

def calculate_monthly_performance_from_records(branch, team_type, year, month):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year+1, 1, 1) - timedelta(days=1)
    else:
        end_date = datetime(year, month+1, 1) - timedelta(days=1)
    
    perf_df = get_team_performance(branch, team_type, start_date, end_date)
    
    if perf_df.empty:
        return None
    
    avg_efficiency = perf_df["efficiency"].mean()
    total_raw = perf_df["raw_water"].sum()
    total_output = perf_df["treated_water"].sum()
    revenue_water = perf_df["revenue_water"].sum()
    backwashing = perf_df["backwashing"].sum()
    non_revenue_pct = perf_df["non_revenue_pct"].mean()
    total_cost = perf_df["expense"].sum()
    total_revenue = perf_df["revenue"].sum()
    total_profit = total_revenue - total_cost
    cost_per_m3 = total_cost / revenue_water if revenue_water > 0 else 0
    avg_quality = perf_df["quality_score"].mean()
    
    return {
        "avg_efficiency": round(avg_efficiency, 1) if not perf_df.empty else 0.0,
        "total_raw": round(total_raw, 1) if not perf_df.empty else 0.0,
        "total_output": round(total_output, 1) if not perf_df.empty else 0.0,
        "revenue_water": round(revenue_water, 1) if not perf_df.empty else 0.0,
        "backwashing": round(backwashing, 1) if not perf_df.empty else 0.0,
        "non_revenue_pct": round(non_revenue_pct, 1) if not perf_df.empty else 0.0,
        "total_cost": round(total_cost, 1) if not perf_df.empty else 0.0,
        "total_revenue": round(total_revenue, 1) if not perf_df.empty else 0.0,
        "total_profit": round(total_profit, 1) if not perf_df.empty else 0.0,
        "cost_per_m3": round(cost_per_m3, 1) if not perf_df.empty else 0.0,
        "compliance_score": round(avg_quality, 1) if not perf_df.empty else 0.0,
        "record_count": len(perf_df)
    }

def calculate_team_rankings(branch, year, month):
    rankings = []
    
    for team in TEAM_TYPES:
        perf = calculate_monthly_performance_from_records(branch, team, year, month)
        if perf and perf["record_count"] > 0:
            rankings.append({
                "team": team,
                "revenue": perf["total_revenue"],
                "efficiency": perf["avg_efficiency"],
                "revenue_water": perf["revenue_water"],
                "cost": perf["cost_per_m3"],
                "compliance": perf["compliance_score"],
                "records": perf["record_count"],
                "raw_water": perf["total_raw"],
                "total_output": perf["total_output"],
                "total_cost": perf["total_cost"],
                "total_profit": perf["total_profit"]
            })
    
    rankings.sort(key=lambda x: x["revenue"], reverse=True)
    
    for i, r in enumerate(rankings, 1):
        r["rank"] = i
    
    return rankings

# ================= AUTOMATIC JAR TEST =================
def automatic_jar_test(flow_m3, raw_ntu, stock_pct):
    if raw_ntu <= 100:
        dose_mgL = 2.0
    elif raw_ntu <= 300:
        dose_mgL = 3.5
    elif raw_ntu <= 600:
        dose_mgL = 6.0
    elif raw_ntu <= 1000:
        dose_mgL = 9.0
    elif raw_ntu <= 2000:
        dose_mgL = 14.0
    elif raw_ntu <= 3000:
        dose_mgL = 22.0
    elif raw_ntu <= 4000:
        dose_mgL = 30.0
    elif raw_ntu <= 5000:
        dose_mgL = 35.0
    else:
        dose_mgL = 45.0
    
    efficiency = 85 + (dose_mgL * 0.65)
    efficiency = min(99.5, efficiency)
    outlet_ntu = raw_ntu * (1 - efficiency/100)
    outlet_ntu = max(0.5, round(outlet_ntu, 1))
    
    stock_conc_gL = stock_pct * 10
    pump_Lh = (flow_m3 * dose_mgL) / stock_conc_gL if stock_conc_gL > 0 else 0
    pump_Lmin = pump_Lh / 60 if pump_Lh > 0 else 0
    
    return {
        'dose_mgL': dose_mgL,
        'outlet_ntu': outlet_ntu,
        'efficiency': round(efficiency, 1),
        'pump_Lh': round(pump_Lh, 1),
        'pump_Lmin': round(pump_Lmin, 1),
        'pump_mLmin': round(pump_Lmin * 1000, 0)
    }

# ================= SHIFT SUMMARY =================
def calculate_shift_summary_for_display(df, shift_type, base_cost):
    if df is None or df.empty:
        return {
            "total_raw_water": 0, "total_treated": 0, "total_expense": 0,
            "total_revenue": 0, "efficiency": 0, "quality_efficiency": 0,
            "revenue_water": 0, "backwashing": 0,
            "record_count": 0
        }
    
    if 'Shift' in df.columns:
        shift_col = 'Shift'
    elif 'shift' in df.columns:
        shift_col = 'shift'
    else:
        shift_col = None
    
    if shift_col:
        shift_df = df[df[shift_col] == shift_type]
    else:
        shift_df = df
    
    if shift_df.empty:
        return {
            "total_raw_water": 0, "total_treated": 0, "total_expense": 0,
            "total_revenue": 0, "efficiency": 0, "quality_efficiency": 0,
            "revenue_water": 0, "backwashing": 0,
            "record_count": 0
        }
    
    if "inflow_m3" in shift_df.columns:
        raw_col = "inflow_m3"
    elif "Inflow_m3" in shift_df.columns:
        raw_col = "Inflow_m3"
    else:
        raw_col = None
    
    if "outflow_m3" in shift_df.columns:
        treated_col = "outflow_m3"
    elif "Outflow_m3" in shift_df.columns:
        treated_col = "Outflow_m3"
    else:
        treated_col = None
    
    if "total_production_cost" in shift_df.columns:
        expense_col = "total_production_cost"
    elif "Total_Production_Cost_RWF" in shift_df.columns:
        expense_col = "Total_Production_Cost_RWF"
    else:
        expense_col = None
    
    if "revenue_water_m3" in shift_df.columns:
        revenue_water_col = "revenue_water_m3"
    elif "Revenue_Water_m3" in shift_df.columns:
        revenue_water_col = "Revenue_Water_m3"
    else:
        revenue_water_col = None
    
    if "backwashing_m3" in shift_df.columns:
        backwashing_col = "backwashing_m3"
    elif "Backwashing_m3" in shift_df.columns:
        backwashing_col = "Backwashing_m3"
    else:
        backwashing_col = None
    
    total_raw = shift_df[raw_col].sum() if raw_col else 0
    total_treated = shift_df[treated_col].sum() if treated_col else 0
    total_expense = shift_df[expense_col].sum() if expense_col else 0
    total_revenue_water = shift_df[revenue_water_col].sum() if revenue_water_col else total_treated * 0.9
    total_backwashing = shift_df[backwashing_col].sum() if backwashing_col else 0
    total_revenue = total_revenue_water * base_cost
    
    if "efficiency" in shift_df.columns:
        avg_eff = shift_df["efficiency"].mean()
    elif "Efficiency_%" in shift_df.columns:
        avg_eff = shift_df["Efficiency_%"].mean()
    elif raw_col and treated_col and total_raw > 0:
        avg_eff = (total_treated / total_raw) * 100
    else:
        avg_eff = 0
    
    return {
        "total_raw_water": round(total_raw, 1),
        "total_treated": round(total_treated, 1),
        "total_expense": round(total_expense, 1),
        "total_revenue": round(total_revenue, 1),
        "efficiency": round(avg_eff, 1),
        "quality_efficiency": round(avg_eff * 0.95, 1),
        "revenue_water": round(total_revenue_water, 1),
        "backwashing": round(total_backwashing, 1),
        "record_count": len(shift_df)
    }

# ================= ENHANCED EXCEL EXPORT FUNCTION (SINGLE ROW PER SHIFT) =================
def export_full_excel_report(branch, year, month, day, filter_level):
    """Export comprehensive Excel report with all meter readings in single row per shift"""
    from io import BytesIO
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Get all data
    all_data = get_records(branch)
    if all_data.empty:
        return None
    
    # Filter data based on selection
    filtered_data = all_data.copy()
    filtered_data["Date"] = pd.to_datetime(filtered_data["Date"])
    
    if filter_level == "Year":
        filtered_data = filtered_data[filtered_data["Date"].dt.year == year]
    elif filter_level == "Month":
        filtered_data = filtered_data[filtered_data["Date"].dt.year == year]
        filtered_data = filtered_data[filtered_data["Date"].dt.month == month]
    elif filter_level == "Day":
        filtered_data = filtered_data[filtered_data["Date"].dt.year == year]
        filtered_data = filtered_data[filtered_data["Date"].dt.month == month]
        filtered_data = filtered_data[filtered_data["Date"].dt.day == day]
    
    if filtered_data.empty:
        return None
    
    output = BytesIO()
    current_thresholds = load_thresholds_for_branch(branch)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ============================================================
        # SHEET 1: METER READINGS (Single row per shift)
        # ============================================================
        shift_col = 'Shift' if 'Shift' in filtered_data.columns else 'shift' if 'shift' in filtered_data.columns else None
        
        if shift_col:
            day_data = filtered_data[filtered_data[shift_col] == "Day"] if not filtered_data.empty else pd.DataFrame()
            night_data = filtered_data[filtered_data[shift_col] == "Night"] if not filtered_data.empty else pd.DataFrame()
        else:
            day_data = filtered_data
            night_data = pd.DataFrame()
        
        # Get all meters
        meters_config = load_meters_config(branch)
        
        # Build column structure
        columns = ["Shift", "Date", "Team"]
        
        # Add meter columns: Previous, Current, Consumption for each meter
        meter_cols = {}
        for meter_type, meters_list in meters_config.items():
            type_label = meter_type.replace("_", " ").title()
            for meter_name in meters_list:
                base_name = f"{type_label} - {meter_name}"
                columns.append(f"{base_name} - Previous")
                columns.append(f"{base_name} - Current")
                columns.append(f"{base_name} - Consumption")
                meter_cols[f"{type_label} - {meter_name}"] = {
                    "type": meter_type,
                    "name": meter_name,
                    "prev_col": f"{base_name} - Previous",
                    "cur_col": f"{base_name} - Current",
                    "cons_col": f"{base_name} - Consumption"
                }
        
        # Add summary columns
        columns.extend([
            "Raw Water Total (m³)",
            "Treated Water Total (m³)",
            "Backwashing Total (m³)",
            "Revenue Water (m³)",
            "Non-Revenue (%)",
            "Production Cost (RWF)",
            "Water Revenue (RWF)",
            "Net Profit (RWF)",
            "Efficiency (%)"
        ])
        
        # Function to get meter reading for specific date
        def get_meter_reading_for_date(date_str, meter_type, meter_name, reading_type="current"):
            readings = get_meter_readings_history(branch, meter_type, meter_name)
            if readings.empty:
                return 0
            
            # Filter by date
            readings = readings[readings["reading_date"] == date_str]
            if readings.empty:
                return 0
            
            # Get the latest reading for that date
            readings = readings.sort_values("recorded_time", ascending=False)
            if reading_type == "current":
                return readings.iloc[0]["reading_value"]
            elif reading_type == "previous":
                # Get previous day's reading
                prev_date = datetime.strptime(date_str, "%Y-%m-%d") - timedelta(days=1)
                prev_readings = get_meter_readings_history(branch, meter_type, meter_name)
                prev_readings = prev_readings[prev_readings["reading_date"] == prev_date.strftime("%Y-%m-%d")]
                if not prev_readings.empty:
                    prev_readings = prev_readings.sort_values("recorded_time", ascending=False)
                    return prev_readings.iloc[0]["reading_value"]
                return 0
            return 0
        
        # Create rows for Day and Night shifts
        rows_data = []
        
        for shift_name, shift_data in [("Day", day_data), ("Night", night_data)]:
            if shift_data.empty:
                # Add empty row
                row = {"Shift": shift_name, "Date": "", "Team": ""}
                for col in columns[3:]:
                    row[col] = 0
                rows_data.append(row)
                continue
            
            # Get date from shift data
            shift_date = shift_data.iloc[0]["date"] if "date" in shift_data.columns else ""
            team = shift_data.iloc[0]["team"] if "team" in shift_data.columns else ""
            
            row = {
                "Shift": shift_name,
                "Date": shift_date,
                "Team": team
            }
            
            # Add meter readings
            for meter_key, meter_info in meter_cols.items():
                meter_type = meter_info["type"]
                meter_name = meter_info["name"]
                
                # Get current reading
                current_val = get_meter_reading_for_date(shift_date, meter_type, meter_name, "current")
                prev_val = get_meter_reading_for_date(shift_date, meter_type, meter_name, "previous")
                consumption = current_val - prev_val if current_val > prev_val else 0
                
                row[meter_info["prev_col"]] = prev_val
                row[meter_info["cur_col"]] = current_val
                row[meter_info["cons_col"]] = consumption
            
            # Add summary values from shift data
            raw_total = shift_data["inflow_m3"].sum() if "inflow_m3" in shift_data.columns else 0
            treated_total = shift_data["outflow_m3"].sum() if "outflow_m3" in shift_data.columns else 0
            backwash_total = shift_data["backwashing_m3"].sum() if "backwashing_m3" in shift_data.columns else 0
            revenue_water = shift_data["revenue_water_m3"].sum() if "revenue_water_m3" in shift_data.columns else 0
            non_rev = shift_data["non_revenue_pct"].mean() if "non_revenue_pct" in shift_data.columns else float(current_thresholds.get("Non_Revenue_Pct", 10.0))
            cost = shift_data["total_production_cost"].sum() if "total_production_cost" in shift_data.columns else 0
            revenue = shift_data["water_revenue"].sum() if "water_revenue" in shift_data.columns else 0
            profit = shift_data["net_profit"].sum() if "net_profit" in shift_data.columns else 0
            efficiency = shift_data["efficiency"].mean() if "efficiency" in shift_data.columns else 0
            
            row["Raw Water Total (m³)"] = round(raw_total, 1)
            row["Treated Water Total (m³)"] = round(treated_total, 1)
            row["Backwashing Total (m³)"] = round(backwash_total, 1)
            row["Revenue Water (m³)"] = round(revenue_water, 1)
            row["Non-Revenue (%)"] = round(non_rev, 1)
            row["Production Cost (RWF)"] = round(cost, 1)
            row["Water Revenue (RWF)"] = round(revenue, 1)
            row["Net Profit (RWF)"] = round(profit, 1)
            row["Efficiency (%)"] = round(efficiency, 1)
            
            rows_data.append(row)
        
        # Create DataFrame
        meter_df = pd.DataFrame(rows_data)
        
        # Reorder columns
        meter_df = meter_df[columns]
        
        # Write to Excel
        meter_df.to_excel(writer, sheet_name="Meter Readings", index=False)
        
        # Format the worksheet
        worksheet = writer.sheets["Meter Readings"]
        
        # Style settings
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
        shift_fill_day = PatternFill(start_color="d4eaf7", end_color="d4eaf7", fill_type="solid")
        shift_fill_night = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styles
        for col in range(1, len(columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Apply styles for data rows
        for row in range(2, len(meter_df) + 2):
            shift_val = worksheet.cell(row=row, column=1).value
            for col in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Color based on shift
                if shift_val == "Day":
                    cell.fill = shift_fill_day
                elif shift_val == "Night":
                    cell.fill = shift_fill_night
                
                # Bold for summary columns
                if col > len(columns) - 9:  # Summary columns
                    cell.font = Font(bold=True)
                
                # Color code consumption values
                if "Consumption" in columns[col-1]:
                    if cell.value and cell.value > 0:
                        cell.font = Font(color="27ae60", bold=True)
        
        # Auto-adjust column widths
        for col in range(1, len(columns) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, len(meter_df) + 2):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 3, 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # ============================================================
        # SHEET 2: DAILY SHIFT SUMMARY
        # ============================================================
        def calc_shift_metrics(df, shift_name):
            if df.empty:
                return {
                    "Shift": shift_name,
                    "Raw Water (m³)": 0,
                    "Treated (m³)": 0,
                    "Revenue Water (m³)": 0,
                    "Backwashing (m³)": 0,
                    "Production Cost (RWF)": 0,
                    "Water Revenue (RWF)": 0,
                    "Net Profit (RWF)": 0,
                    "Efficiency (%)": 0,
                    "Non-Revenue (%)": float(current_thresholds.get("Non_Revenue_Pct", 10.0)),
                    "Records Count": 0
                }
            
            raw = df['inflow_m3'].sum() if 'inflow_m3' in df.columns else 0
            treated = df['outflow_m3'].sum() if 'outflow_m3' in df.columns else 0
            revenue_water = df['revenue_water_m3'].sum() if 'revenue_water_m3' in df.columns else 0
            backwashing = df['backwashing_m3'].sum() if 'backwashing_m3' in df.columns else 0
            cost = df['total_production_cost'].sum() if 'total_production_cost' in df.columns else 0
            revenue = df['water_revenue'].sum() if 'water_revenue' in df.columns else 0
            eff = df['efficiency'].mean() if 'efficiency' in df.columns else (treated/raw*100 if raw > 0 else 0)
            
            return {
                "Shift": shift_name,
                "Raw Water (m³)": round(raw, 1),
                "Treated (m³)": round(treated, 1),
                "Revenue Water (m³)": round(revenue_water, 1),
                "Backwashing (m³)": round(backwashing, 1),
                "Production Cost (RWF)": round(cost, 1),
                "Water Revenue (RWF)": round(revenue, 1),
                "Net Profit (RWF)": round(revenue - cost, 1),
                "Efficiency (%)": round(eff, 1),
                "Non-Revenue (%)": float(current_thresholds.get("Non_Revenue_Pct", 10.0)),
                "Records Count": len(df)
            }
        
        day_summary = calc_shift_metrics(day_data, "Day")
        night_summary = calc_shift_metrics(night_data, "Night")
        
        total_raw = day_summary["Raw Water (m³)"] + night_summary["Raw Water (m³)"]
        total_treated = day_summary["Treated (m³)"] + night_summary["Treated (m³)"]
        total_revenue_water = day_summary["Revenue Water (m³)"] + night_summary["Revenue Water (m³)"]
        total_backwashing = day_summary["Backwashing (m³)"] + night_summary["Backwashing (m³)"]
        total_cost = day_summary["Production Cost (RWF)"] + night_summary["Production Cost (RWF)"]
        total_revenue = day_summary["Water Revenue (RWF)"] + night_summary["Water Revenue (RWF)"]
        total_profit = total_revenue - total_cost
        avg_eff = (day_summary["Efficiency (%)"] + night_summary["Efficiency (%)"]) / 2 if (day_summary["Efficiency (%)"] > 0 or night_summary["Efficiency (%)"] > 0) else 0
        
        daily_total = {
            "Shift": "DAILY TOTAL",
            "Raw Water (m³)": round(total_raw, 1),
            "Treated (m³)": round(total_treated, 1),
            "Revenue Water (m³)": round(total_revenue_water, 1),
            "Backwashing (m³)": round(total_backwashing, 1),
            "Production Cost (RWF)": round(total_cost, 1),
            "Water Revenue (RWF)": round(total_revenue, 1),
            "Net Profit (RWF)": round(total_profit, 1),
            "Efficiency (%)": round(avg_eff, 1),
            "Non-Revenue (%)": float(current_thresholds.get("Non_Revenue_Pct", 10.0)),
            "Records Count": day_summary["Records Count"] + night_summary["Records Count"]
        }
        
        shift_summary_df = pd.DataFrame([day_summary, night_summary, daily_total])
        shift_summary_df.to_excel(writer, sheet_name="Shift Summary", index=False)
        
        # Format shift summary
        ws_shift = writer.sheets["Shift Summary"]
        for col in range(1, len(shift_summary_df.columns) + 1):
            cell = ws_shift.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for row in range(2, len(shift_summary_df) + 2):
            for col in range(1, len(shift_summary_df.columns) + 1):
                cell = ws_shift.cell(row=row, column=col)
                cell.border = border
                if row == len(shift_summary_df) + 1:
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                    cell.font = Font(bold=True)
        
        for col in range(1, len(shift_summary_df.columns) + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(1, len(shift_summary_df) + 2):
                val = ws_shift.cell(row=row, column=col).value
                if val:
                    max_length = max(max_length, len(str(val)))
            ws_shift.column_dimensions[col_letter].width = min(max_length + 3, 30)
        
        # ============================================================
        # SHEET 3: CHEMICALS & CONSUMPTION
        # ============================================================
        price_lime = float(current_thresholds.get("Price_Lime", 450.0))
        price_fl = float(current_thresholds.get("Price_FL4440", 2500.0))
        price_nacl = float(current_thresholds.get("Price_NaCl", 300.0))
        price_elec = float(current_thresholds.get("Price_Electricity", 148.0))
        
        chemicals_data = []
        
        if not filtered_data.empty:
            for _, row in filtered_data.iterrows():
                treated = row.get('outflow_m3', 0)
                lime_kg = treated * 0.001
                fl4440_l = treated * 0.002
                nacl_kg = treated * 0.0005
                elec_kwh = treated * 0.1
                
                lime_cost = lime_kg * price_lime
                fl_cost = fl4440_l * price_fl
                nacl_cost = nacl_kg * price_nacl
                elec_cost = elec_kwh * price_elec
                
                chemicals_data.append({
                    "Date": row.get('date', ''),
                    "Shift": row.get('shift', ''),
                    "Water Treated (m³)": round(treated, 1),
                    "Lime (kg)": round(lime_kg, 2),
                    "FL4440 (L)": round(fl4440_l, 2),
                    "NaCl (kg)": round(nacl_kg, 2),
                    "Electricity (kWh)": round(elec_kwh, 2),
                    "Lime Cost (RWF)": round(lime_cost, 1),
                    "FL4440 Cost (RWF)": round(fl_cost, 1),
                    "NaCl Cost (RWF)": round(nacl_cost, 1),
                    "Electricity Cost (RWF)": round(elec_cost, 1),
                    "Total Cost (RWF)": round(lime_cost + fl_cost + nacl_cost + elec_cost, 1)
                })
        
        if chemicals_data:
            chemicals_df = pd.DataFrame(chemicals_data)
            chemicals_df.to_excel(writer, sheet_name="Chemicals & Energy", index=False)
            
            ws_chem = writer.sheets["Chemicals & Energy"]
            for col in range(1, len(chemicals_df.columns) + 1):
                cell = ws_chem.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            for col in range(1, len(chemicals_df.columns) + 1):
                max_length = 0
                col_letter = get_column_letter(col)
                for row in range(1, len(chemicals_df) + 2):
                    val = ws_chem.cell(row=row, column=col).value
                    if val:
                        max_length = max(max_length, len(str(val)))
                ws_chem.column_dimensions[col_letter].width = min(max_length + 3, 30)
        
        # ============================================================
        # SHEET 4: COMPLETE DETAILED RECORDS
        # ============================================================
        if not filtered_data.empty:
            columns_to_export = [
                'date', 'shift', 'team', 'inflow_m3', 'outflow_m3', 
                'backwashing_m3', 'revenue_water_m3', 'non_revenue_pct',
                'total_production_cost', 'water_revenue', 'net_profit', 
                'efficiency', 'ph_level', 'turbidity_ntu', 'temperature',
                'chlorine_mgL', 'nitrate_mgL', 'recorded_by'
            ]
            
            existing_cols = [col for col in columns_to_export if col in filtered_data.columns]
            detailed_df = filtered_data[existing_cols].copy()
            
            rename_map = {
                'date': 'Date',
                'shift': 'Shift',
                'team': 'Team',
                'inflow_m3': 'Raw Water (m³)',
                'outflow_m3': 'Treated Water (m³)',
                'backwashing_m3': 'Backwashing (m³)',
                'revenue_water_m3': 'Revenue Water (m³)',
                'non_revenue_pct': 'Non-Revenue (%)',
                'total_production_cost': 'Production Cost (RWF)',
                'water_revenue': 'Water Revenue (RWF)',
                'net_profit': 'Net Profit (RWF)',
                'efficiency': 'Efficiency (%)',
                'ph_level': 'pH Level',
                'turbidity_ntu': 'Turbidity (NTU)',
                'temperature': 'Temperature (°C)',
                'chlorine_mgL': 'Chlorine (mg/L)',
                'nitrate_mgL': 'Nitrate (mg/L)',
                'recorded_by': 'Recorded By'
            }
            
            detailed_df = detailed_df.rename(columns=rename_map)
            detailed_df.to_excel(writer, sheet_name="Detailed Records", index=False)
            
            ws_detail = writer.sheets["Detailed Records"]
            for col in range(1, len(detailed_df.columns) + 1):
                cell = ws_detail.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            for col in range(1, len(detailed_df.columns) + 1):
                max_length = 0
                col_letter = get_column_letter(col)
                for row in range(1, len(detailed_df) + 2):
                    val = ws_detail.cell(row=row, column=col).value
                    if val:
                        max_length = max(max_length, len(str(val)))
                ws_detail.column_dimensions[col_letter].width = min(max_length + 3, 30)
        
        # ============================================================
        # SHEET 5: WATER QUALITY DATA
        # ============================================================
        wq_data = get_water_quality_records(branch, limit=500)
        if not wq_data.empty:
            wq_export = wq_data[['datetime', 'flow_m3', 'raw_ntu', 'outlet_ntu', 
                                 'efficiency', 'pump_Lh', 'raw_ph', 'raw_temperature',
                                 'raw_chlorine', 'raw_nitrate']].copy()
            wq_export.columns = ['DateTime', 'Flow (m³/h)', 'Raw Turbidity (NTU)', 
                                'Outlet Turbidity (NTU)', 'Efficiency (%)', 
                                'Pump (L/h)', 'pH', 'Temperature (°C)',
                                'Chlorine (mg/L)', 'Nitrate (mg/L)']
            wq_export.to_excel(writer, sheet_name="Water Quality", index=False)
            
            ws_wq = writer.sheets["Water Quality"]
            for col in range(1, len(wq_export.columns) + 1):
                cell = ws_wq.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            for col in range(1, len(wq_export.columns) + 1):
                max_length = 0
                col_letter = get_column_letter(col)
                for row in range(1, len(wq_export) + 2):
                    val = ws_wq.cell(row=row, column=col).value
                    if val:
                        max_length = max(max_length, len(str(val)))
                ws_wq.column_dimensions[col_letter].width = min(max_length + 3, 30)
    
    output.seek(0)
    return output

# ================= FORMAT NUMBERS =================
def format_number(num):
    if num >= 1_000_000:
        return f"{num:,.0f}"
    elif num >= 1_000:
        return f"{num:,.0f}"
    else:
        return f"{num:.1f}"

def get_date_range_display(filtered_data):
    if filtered_data.empty:
        return "No Data", ""
    
    shift_col = None
    if 'Shift' in filtered_data.columns:
        shift_col = 'Shift'
    elif 'shift' in filtered_data.columns:
        shift_col = 'shift'
    
    if shift_col is None:
        return "No Data", ""
    
    has_day = not filtered_data[filtered_data[shift_col] == 'Day'].empty
    has_night = not filtered_data[filtered_data[shift_col] == 'Night'].empty
    
    if has_day and has_night:
        day_date = filtered_data[filtered_data[shift_col] == 'Day']['Date'].dt.date.min()
        night_date = filtered_data[filtered_data[shift_col] == 'Night']['Date'].dt.date.max()
        
        if day_date == night_date:
            display_info = day_date.strftime("%Y/%m/%d")
            display_sub = ""
        else:
            display_info = f"{day_date.strftime('%Y/%m/%d')} - {night_date.strftime('%Y/%m/%d')}"
            display_sub = ""
    elif has_day:
        day_date = filtered_data[filtered_data[shift_col] == 'Day']['Date'].dt.date.min()
        display_info = day_date.strftime("%Y/%m/%d")
        display_sub = ""
    elif has_night:
        night_date = filtered_data[filtered_data[shift_col] == 'Night']['Date'].dt.date.max()
        display_info = night_date.strftime("%Y/%m/%d")
        display_sub = ""
    else:
        display_info = "No Data"
        display_sub = ""
    
    return display_info, display_sub

# ================= INITIALIZE SESSION STATE =================
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False

if 'user_role' not in st.session_state:
    st.session_state.user_role = None

if 'user_branch' not in st.session_state:
    st.session_state.user_branch = "NZOVE"

if 'show_reset' not in st.session_state:
    st.session_state.show_reset = False

if 'reset_user' not in st.session_state:
    st.session_state.reset_user = None

if 'login_attempts' not in st.session_state:
    st.session_state.login_attempts = 0

if 'menu_option' not in st.session_state:
    st.session_state.menu_option = "Dashboard"

if 'selected_date' not in st.session_state:
    st.session_state.selected_date = date.today()

if 'selected_year' not in st.session_state:
    st.session_state.selected_year = 2026

if 'selected_month' not in st.session_state:
    st.session_state.selected_month = datetime.now().month

if 'selected_day' not in st.session_state:
    st.session_state.selected_day = datetime.now().day

if 'filter_level' not in st.session_state:
    st.session_state.filter_level = "Year"

if 'data_refresh_counter' not in st.session_state:
    st.session_state.data_refresh_counter = 0

current_year = datetime.now().year

# ================= CSS STYLES =================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

* {
    font-family: 'Inter', sans-serif;
}

.stApp {
    background: linear-gradient(135deg, #87CEEB 0%, #87CEEB 25%, #87CEEB 50%, #87CEEB 75%, #87CEEB 100%);
    background-attachment: fixed;
    color: #1a1a2e;
}

.big-title {
    text-align: center;
    background: linear-gradient(135deg, #1e3c72, #2a5298, #1e3c72, #2a5298);
    background-size: 300% 300%;
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    font-size: 38px;
    font-weight: 800;
    padding: 20px;
    animation: gradientShift 4s ease infinite;
}

@keyframes gradientShift {
    0% { background-position: 0% 50%; }
    50% { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}

.branch-header {
    background: linear-gradient(135deg, rgba(255,255,255,0.95), rgba(240,248,255,0.95));
    padding: 15px 25px;
    border-radius: 24px;
    text-align: center;
    margin-bottom: 25px;
    border: 1px solid rgba(30,60,114,0.3);
    box-shadow: 0 8px 32px rgba(0,0,0,0.1);
    backdrop-filter: blur(10px);
}

.branch-header span {
    font-size: 24px;
    font-weight: 700;
    background: linear-gradient(135deg, #1e3c72, #2a5298);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
}

.card {
    background: linear-gradient(135deg, rgba(255,255,255,0.95), rgba(240,248,255,0.95));
    padding: 12px 8px;
    border-radius: 20px;
    text-align: center;
    border: 1px solid rgba(30,60,114,0.25);
    transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
    backdrop-filter: blur(12px);
    position: relative;
    overflow: hidden;
    min-height: 80px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    word-wrap: break-word;
    overflow-wrap: break-word;
    white-space: normal;
}

.card::before {
    content: '';
    position: absolute;
    top: 0;
    left: -100%;
    width: 100%;
    height: 100%;
    background: linear-gradient(90deg, transparent, rgba(30,60,114,0.1), transparent);
    transition: left 0.6s;
}

.card:hover::before {
    left: 100%;
}

.card:hover {
    transform: translateY(-3px) scale(1.01);
    border-color: #1e3c72;
    box-shadow: 0 10px 25px rgba(30,60,114,0.15);
}

.card-label {
    font-size: 10px;
    font-weight: 600;
    color: #1e3c72;
    letter-spacing: 0.3px;
    margin-bottom: 5px;
    text-transform: uppercase;
}

.card-val {
    font-size: 20px;
    font-weight: 800;
    background: linear-gradient(135deg, #1e3c72, #2a5298);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    line-height: 1.2;
    word-wrap: break-word;
    overflow-wrap: break-word;
    white-space: normal;
}

.card-val-large {
    font-size: 22px;
    font-weight: 800;
    background: linear-gradient(135deg, #1e3c72, #e67e22);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    word-wrap: break-word;
    overflow-wrap: break-word;
    white-space: normal;
}

.card-unit {
    font-size: 8px;
    color: #64748b;
    margin-top: 3px;
}

.graph-container {
    background: rgba(255,255,255,0.9);
    border-radius: 20px;
    padding: 15px;
    margin-bottom: 20px;
    border: 1px solid rgba(30,60,114,0.2);
    backdrop-filter: blur(10px);
    transition: all 0.3s;
}

.graph-container:hover {
    border-color: #1e3c72;
    box-shadow: 0 8px 32px rgba(30,60,114,0.1);
}

.graph-title {
    font-size: 15px;
    font-weight: 700;
    margin-bottom: 12px;
    padding-left: 12px;
    border-left: 4px solid #1e3c72;
    color: #1e3c72;
    display: flex;
    align-items: center;
    gap: 8px;
}

.graph-comment {
    background: linear-gradient(135deg, rgba(240,248,255,0.8), rgba(255,255,255,0.8));
    border-left: 4px solid #1e3c72;
    padding: 10px 14px;
    border-radius: 14px;
    margin-top: 12px;
    font-size: 11px;
    line-height: 1.5;
    color: #1a1a2e;
}

.custom-divider {
    height: 2px;
    background: linear-gradient(90deg, transparent, #1e3c72, #2a5298, #1e3c72, transparent);
    margin: 15px 0;
    animation: shimmer 3s infinite;
}

@keyframes shimmer {
    0% { background-position: -200% 0; }
    100% { background-position: 200% 0; }
}

.subtotal-container {
    background: linear-gradient(135deg, #ffffff, #f0f8ff);
    padding: 8px 14px;
    border-radius: 12px;
    border-left: 4px solid #27ae60;
    margin: 8px 0;
    font-size: 12px;
    transition: all 0.3s;
}

.subtotal-container:hover {
    transform: translateX(5px);
    border-left-color: #1e3c72;
}

.team-card {
    background: linear-gradient(135deg, rgba(255,255,255,0.95), rgba(240,248,255,0.95));
    border-radius: 16px;
    padding: 15px;
    border: 1px solid rgba(30,60,114,0.3);
    margin-top: 10px;
    transition: all 0.3s;
}

.team-card:hover {
    border-color: #1e3c72;
    box-shadow: 0 5px 20px rgba(30,60,114,0.1);
}

.settings-section {
    background: rgba(255,255,255,0.95);
    border-radius: 16px;
    padding: 15px;
    margin-bottom: 15px;
    border: 1px solid rgba(30,60,114,0.2);
}

.settings-section h4 {
    color: #1e3c72;
    font-weight: 700;
    margin-bottom: 12px;
}

.water-quality-card {
    background: linear-gradient(135deg, #ffffff, #f0f8ff);
    border-radius: 20px;
    padding: 20px;
    margin-bottom: 20px;
    border: 1px solid rgba(30,60,114,0.3);
    box-shadow: 0 4px 15px rgba(0,0,0,0.05);
}

.stButton>button {
    background: linear-gradient(135deg, #1e3c72, #2a5298);
    color: white;
    font-weight: 600;
    border: none;
    border-radius: 12px;
    padding: 8px 18px;
    font-size: 13px;
    transition: all 0.3s;
}

.stButton>button:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(30,60,114,0.4);
}

.sidebar-section-title {
    font-size: 14px;
    font-weight: 700;
    color: #000000 !important;
    margin-top: 15px;
    margin-bottom: 10px;
    padding-left: 5px;
    letter-spacing: 0.5px;
}

.status-dot {
    display: inline-block;
    width: 8px;
    height: 8px;
    border-radius: 50%;
    animation: pulse 2s infinite;
    margin-right: 5px;
}

@keyframes pulse {
    0%, 100% { opacity: 1; transform: scale(1); }
    50% { opacity: 0.5; transform: scale(1.2); }
}

.status-dot.green { background: #27ae60; box-shadow: 0 0 8px #27ae60; }
.status-dot.yellow { background: #f39c12; box-shadow: 0 0 8px #f39c12; }
.status-dot.red { background: #e74c3c; box-shadow: 0 0 8px #e74c3c; }

[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #e8f4f8, #d4eaf7) !important;
    border-right: 1px solid rgba(30,60,114,0.15) !important;
    backdrop-filter: blur(10px) !important;
}

[data-testid="stSidebar"] .stMarkdown,
[data-testid="stSidebar"] .stMarkdown p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div {
    color: #000000 !important;
}

.streamlit-expanderHeader {
    background: linear-gradient(135deg, #ffffff, #f0f8ff) !important;
    border-radius: 12px !important;
    font-weight: 600 !important;
    color: #000000 !important;
    border: 1px solid rgba(30,60,114,0.2) !important;
}

.streamlit-expanderHeader:hover {
    border-color: #1e3c72 !important;
}

.dataframe, [data-testid="stDataFrame"], .stDataFrame {
    background: #ffffff !important;
    border-radius: 16px !important;
    border: 1px solid #d4eaf7 !important;
    font-size: 12px;
    width: 100% !important;
    overflow-x: auto !important;
    display: block !important;
}

.shift-box {
    background: rgba(255,255,255,0.9);
    border-radius: 16px;
    padding: 10px;
    text-align: center;
    border: 1px solid #1e3c72;
    margin: 5px;
    word-wrap: break-word;
    overflow-wrap: break-word;
    white-space: normal;
}
</style>
""", unsafe_allow_html=True)

def tooltip(text):
    return f"<span title='{text}' style='cursor:help;color:#1e3c72;'>ⓘ</span>"

# ================= SIDEBAR =================
st.sidebar.title("Nzove Monitor")

if not st.session_state.logged_in:
    if st.session_state.login_attempts >= 3:
        st.sidebar.error("Too many failed attempts.")
        st.stop()
    
    menu = ["Login", "SignUp"]
    choice = st.sidebar.selectbox("Menu", menu, key="menu_select")

    if choice == "SignUp":
        st.sidebar.markdown(f"**Create Account** {tooltip('Username unique, password 6+ chars with uppercase and number')}", unsafe_allow_html=True)
        new_user = st.sidebar.text_input("Username", key="signup_username")
        new_password = st.sidebar.text_input("Password", type="password", key="signup_password")
        confirm_password = st.sidebar.text_input("Confirm Password", type="password", key="signup_confirm")
        role_choice = st.sidebar.selectbox("Role", ["User", "Admin"], key="signup_role")
        branch_choice = st.sidebar.selectbox("Assign Branch", ["NZOVE"], key="signup_branch")
        
        if st.sidebar.button("Create Account", use_container_width=True, key="signup_btn"):
            if new_user.strip() and new_password == confirm_password:
                if len(new_password) >= 6 and any(c.isupper() for c in new_password) and any(c.isdigit() for c in new_password):
                    if save_user(new_user, new_password, role_choice, branch_choice):
                        st.sidebar.success("Account created! Please login.")
                        st.rerun()
                    else:
                        st.sidebar.error("Username already exists!")
                else:
                    st.sidebar.error("Password must be 6+ chars, include uppercase and number")
            else:
                st.sidebar.error("Passwords do not match or username empty")
    else:
        st.sidebar.markdown(f"**Login** {tooltip('Default: admin / admin123')}", unsafe_allow_html=True)
        username = st.sidebar.text_input("Username", key="login_username")
        password = st.sidebar.text_input("Password", type="password", key="login_password")
        if st.sidebar.button("Login", use_container_width=True, key="login_btn"):
            users = load_users()
            if username in users and check_hashes(password, users[username]["password"]):
                st.session_state.logged_in = True
                st.session_state.current_user = username
                st.session_state.user_role = users[username]["role"]
                st.session_state.user_branch = users[username].get("branch", "NZOVE")
                st.session_state.login_attempts = 0
                st.rerun()
            else:
                st.session_state.login_attempts += 1
                st.sidebar.error(f"Invalid credentials! {3 - st.session_state.login_attempts} attempts remaining")

# ================= MAIN SYSTEM =================
if st.session_state.logged_in:
    st.markdown("<h1 class='big-title'>NZOVE WATER TREATMENT PLANT</h1>", unsafe_allow_html=True)

    selected_branch = "NZOVE"
    current_thresholds = load_thresholds_for_branch(selected_branch)
    active_global_cost = load_global_base_cost()
    non_revenue_pct = float(current_thresholds.get("Non_Revenue_Pct", 10.0))
    
    # ================= SIDEBAR - SECTION STRUCTURE =================
    # SECTION 1: GENERAL SETTINGS (Main section)
    st.sidebar.markdown("<div class='sidebar-section-title'>⚙️ GENERAL SETTINGS</div>", unsafe_allow_html=True)
    with st.sidebar.expander("General Settings", expanded=True):
        st.markdown("**Account**")
        st.markdown(f"**User:** `{st.session_state.current_user}`")
        st.markdown(f"**Role:** `{st.session_state.user_role}`")
        st.markdown(f"**Branch:** `{st.session_state.user_branch}`")
        
        st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
        
        st.markdown("**Change Password**")
        old_pwd = st.text_input("Current Password", type="password", key="change_old_pwd")
        new_pwd = st.text_input("New Password", type="password", key="change_new_pwd")
        confirm_pwd = st.text_input("Confirm Password", type="password", key="change_confirm_pwd")
        if st.button("Update Password", use_container_width=True, key="change_pwd_btn"):
            if new_pwd == confirm_pwd and len(new_pwd) >= 6 and any(c.isupper() for c in new_pwd) and any(c.isdigit() for c in new_pwd):
                if change_user_password(st.session_state.current_user, old_pwd, new_pwd):
                    st.success("Password changed!")
                    st.rerun()
                else:
                    st.error("Incorrect current password")
            else:
                st.error("Password must be 6+ chars, include uppercase and number")
        
        st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
        
        st.markdown("**Security Question**")
        sec_q = get_security_question(st.session_state.current_user)
        if sec_q is None:
            st.warning("No security question set")
            question = st.selectbox("Select Question", ["Mother's maiden name?", "First pet's name?", "Favorite book?", "Birth city?"], key="security_question_select")
            answer = st.text_input("Your Answer", type="password", key="security_answer")
            if st.button("Save Security Question", use_container_width=True, key="save_security_btn"):
                if answer.strip():
                    set_security_question(st.session_state.current_user, question, answer)
                    st.success("Security question saved!")
                    st.rerun()
        else:
            st.info(f"Question: {sec_q}")
        
        st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
        
        st.markdown("**Forgot Password**")
        forgot_user = st.text_input("Username for reset", key="forgot_username")
        if st.button("Request Reset", use_container_width=True, key="reset_request_btn"):
            users = load_users()
            if forgot_user in users:
                sq = get_security_question(forgot_user)
                if sq:
                    st.session_state.reset_user = forgot_user
                    st.session_state.show_reset = True
                    st.info(f"Question: {sq}")
                else:
                    st.error("No security question set")
            else:
                st.error("User not found")
        
        if st.session_state.show_reset:
            answer = st.text_input("Answer", type="password", key="reset_answer")
            new_pass = st.text_input("New Password", type="password", key="reset_new_pass")
            if st.button("Reset Password", use_container_width=True, key="reset_confirm_btn"):
                if verify_security_answer(st.session_state.reset_user, answer):
                    if reset_password_via_security(st.session_state.reset_user, new_pass):
                        st.success("Password reset!")
                        st.session_state.show_reset = False
                        st.rerun()
                    else:
                        st.error("Failed to reset")
                else:
                    st.error("Incorrect answer")
        
        st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
        
        st.markdown("**System**")
        if st.button("Logout", use_container_width=True, key="logout_btn"):
            st.session_state.logged_in = False
            st.cache_data.clear()
            st.rerun()
    
    # SECTION 2: USER MANAGEMENT (Admin only - subset)
    if st.session_state.user_role == "Admin":
        st.sidebar.markdown("<div class='sidebar-section-title'>👥 USER MANAGEMENT</div>", unsafe_allow_html=True)
        with st.sidebar.expander("Manage Users", expanded=False):
            users_list = get_all_users()
            for user in users_list:
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(f"**{user['username']}** - {user['role']}")
                with col2:
                    if user['username'] != "admin" and st.button("Delete", key=f"del_{user['username']}"):
                        if delete_user(user['username']):
                            st.success(f"User deleted!")
                            st.rerun()
            
            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
            st.markdown("**Create User**")
            new_user_admin = st.text_input("Username", key="admin_new_user")
            new_pass_admin = st.text_input("Password", type="password", key="admin_new_pass")
            new_role_admin = st.selectbox("Role", ["User", "Admin"], key="admin_new_role")
            if st.button("Create User", use_container_width=True, key="admin_create_btn"):
                if new_user_admin.strip() and len(new_pass_admin) >= 6:
                    if save_user(new_user_admin, new_pass_admin, new_role_admin, "NZOVE"):
                        st.success(f"User created!")
                        st.rerun()
                    else:
                        st.error("Username already exists!")
                else:
                    st.error("Username and password required")
    
    # SECTION 3: PLANT SETTINGS (subset)
    st.sidebar.markdown("<div class='sidebar-section-title'>🏭 PLANT SETTINGS</div>", unsafe_allow_html=True)
    with st.sidebar.expander("Plant Configuration", expanded=False):
        st.markdown("**Team Information**")
        selected_team_type = st.selectbox("Select Team Type", TEAM_TYPES, key="team_type_select")
        existing_team = get_team_info(selected_branch, selected_team_type)
        
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            team_leader = st.text_input(f"Team {selected_team_type} Leader", value=existing_team.get("team_leader", "") if existing_team else "", key=f"team_leader_{selected_team_type}")
        with col_t2:
            contact = st.text_input(f"Team {selected_team_type} Contact", value=existing_team.get("contact", "") if existing_team else "", key=f"team_contact_{selected_team_type}")
        
        st.markdown("**Team Members**")
        members = st.text_area(
            f"Team {selected_team_type} Members (one per line)",
            value=existing_team.get("members", "") if existing_team else "",
            height=80,
            placeholder="Enter team members names, one per line",
            key=f"team_members_{selected_team_type}"
        )
        
        if st.button(f"Save Team {selected_team_type}", use_container_width=True, key=f"save_team_btn_{selected_team_type}"):
            if team_leader.strip():
                save_team_info(selected_branch, selected_team_type, team_leader, members, "", contact, st.session_state.current_user)
                st.success(f"Team {selected_team_type} saved!")
                st.rerun()
            else:
                st.error("Team Leader name required")
        
        st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
        st.markdown("**All Teams**")
        for team_type in TEAM_TYPES:
            team_data = get_team_info(selected_branch, team_type)
            if team_data:
                st.markdown(f"""
                <div class='team-card'>
                    <b>Team {team_type}</b>: {team_data.get('team_leader', 'N/A')} | {team_data.get('contact', 'N/A')}
                    <br><small><b>Members:</b> {team_data.get('members', 'None')}</small>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"<div class='team-card'><b>Team {team_type}</b>: No info</div>", unsafe_allow_html=True)
        
        if st.session_state.user_role == "Admin":
            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
            st.markdown("**Thresholds & Prices**")
            b_ph_min = st.number_input("pH Min", value=float(current_thresholds.get("pH_min", 6.5)), step=0.1, key="b_ph_min")
            b_ph_max = st.number_input("pH Max", value=float(current_thresholds.get("pH_max", 8.5)), step=0.1, key="b_ph_max")
            b_turb = st.number_input("Turbidity Max (NTU)", value=float(current_thresholds.get("Turbidity_max", 5.0)), step=0.1, key="b_turb")
            b_eff = st.number_input("Min Efficiency (%)", value=float(current_thresholds.get("Efficiency_min", 75.0)), step=1.0, key="b_eff")
            b_non_rev = st.number_input("Non-Revenue (%)", value=float(current_thresholds.get("Non_Revenue_Pct", 10.0)), step=0.5, key="b_non_rev")
            b_cost = st.number_input("Base Price (RWF/m³)", value=float(current_thresholds.get("Base_Cost", 768.0)), key="b_cost")
            
            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
            st.markdown("**Chemical & Energy Prices**")
            
            col_p1, col_p2 = st.columns(2)
            with col_p1:
                price_lime = st.number_input("Lime Price (RWF/kg)", value=float(current_thresholds.get("Price_Lime", 450.0)), step=10.0, key="price_lime")
                price_fl = st.number_input("FL4440 Price (RWF/L)", value=float(current_thresholds.get("Price_FL4440", 2500.0)), step=50.0, key="price_fl")
            with col_p2:
                price_nacl = st.number_input("NaCl Price (RWF/kg)", value=float(current_thresholds.get("Price_NaCl", 300.0)), step=10.0, key="price_nacl")
                price_elec = st.number_input("Electricity Price (RWF/kWh)", value=float(current_thresholds.get("Price_Electricity", 148.0)), step=1.0, key="price_elec")
            
            if st.button(f"Save Config", use_container_width=True, key="save_branch_config"):
                updated = {
                    "pH_min": b_ph_min, "pH_max": b_ph_max, "Turbidity_max": b_turb,
                    "Efficiency_min": b_eff, "Non_Revenue_Pct": b_non_rev, "Base_Cost": b_cost,
                    "Price_Lime": price_lime, "Price_FL4440": price_fl,
                    "Price_NaCl": price_nacl, "Price_Electricity": price_elec
                }
                save_thresholds_for_branch(selected_branch, updated)
                st.success("✅ Config saved!")
                st.rerun()
    
    # SECTION 4: METER MANAGEMENT (Admin only - subset)
    if st.session_state.user_role == "Admin":
        st.sidebar.markdown("<div class='sidebar-section-title'>⚙️ METER MANAGEMENT</div>", unsafe_allow_html=True)
        with st.sidebar.expander("Manage Meters", expanded=False):
            st.markdown("**Current Meters**")
            current_meters = get_all_meters(selected_branch)
            
            for meter_type in ["raw_water", "treated_water", "backwashing", "electricity"]:
                type_label = meter_type.replace("_", " ").title()
                st.markdown(f"**{type_label}**")
                for meter in current_meters.get(meter_type, []):
                    col1, col2, col3 = st.columns([3, 1, 1])
                    with col1:
                        st.write(f"• {meter}")
                    with col2:
                        if st.button("Edit", key=f"edit_{meter_type}_{meter}"):
                            st.session_state.edit_meter_type = meter_type
                            st.session_state.edit_meter_name = meter
                    with col3:
                        if st.button("Remove", key=f"remove_{meter_type}_{meter}"):
                            if remove_meter(selected_branch, meter_type, meter):
                                st.success(f"Removed {meter}")
                                st.rerun()
                
                st.markdown("---")
            
            st.markdown("**Add New Meter**")
            new_meter_type = st.selectbox("Meter Type", ["raw_water", "treated_water", "backwashing", "electricity"], key="new_meter_type")
            new_meter_name = st.text_input("Meter Name", key="new_meter_name")
            if st.button("Add Meter", use_container_width=True, key="add_meter_btn"):
                if new_meter_name.strip():
                    if add_meter(selected_branch, new_meter_type, new_meter_name):
                        st.success(f"Added {new_meter_name}")
                        st.rerun()
                    else:
                        st.error("Meter already exists!")
                else:
                    st.error("Meter name required")
            
            if hasattr(st.session_state, 'edit_meter_name') and st.session_state.edit_meter_name:
                st.markdown("---")
                st.markdown(f"**Editing: {st.session_state.edit_meter_name}**")
                new_name = st.text_input("New Name", value=st.session_state.edit_meter_name, key="edit_meter_new_name")
                if st.button("Save Edit", use_container_width=True, key="save_edit_meter_btn"):
                    if new_name.strip():
                        if edit_meter(selected_branch, st.session_state.edit_meter_type, st.session_state.edit_meter_name, new_name):
                            st.success(f"Updated to {new_name}")
                            st.session_state.edit_meter_name = None
                            st.rerun()
                        else:
                            st.error("Edit failed")
                    else:
                        st.error("Name required")
                if st.button("Cancel Edit", use_container_width=True, key="cancel_edit_meter_btn"):
                    st.session_state.edit_meter_name = None
                    st.rerun()
    
    # SECTION 5: DANGER ZONE (Admin only - subset)
    if st.session_state.user_role == "Admin":
        st.sidebar.markdown("<div class='sidebar-section-title'>⚠️ DANGER ZONE</div>", unsafe_allow_html=True)
        with st.sidebar.expander("🗑️ DELETE ALL DATA", expanded=False):
            st.error("⚠️⚠️⚠️ THIS WILL DELETE ALL RECORDS PERMANENTLY! ⚠️⚠️⚠️")
            st.warning("All historical data will be erased. System will start with ZERO records.")
            st.info("Type 'DELETE ALL' in the box below to confirm")
            confirm_delete_all = st.text_input("Confirm with 'DELETE ALL'", key="confirm_delete_all", type="password")
            
            col_del1, col_del2 = st.columns(2)
            with col_del1:
                if st.button("🗑️ DELETE ALL", use_container_width=True, key="delete_all_btn"):
                    if confirm_delete_all == "DELETE ALL":
                        delete_all_records()
                        st.success("✅ ALL RECORDS DELETED!")
                        st.info("📝 System is now empty. Start adding new records.")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error("❌ Please type 'DELETE ALL' to confirm")
            with col_del2:
                if st.button("🔄 Cancel", use_container_width=True, key="cancel_delete_btn"):
                    st.info("Operation cancelled")
    
    # SECTION 6: TIME FILTER (subset)
    st.sidebar.markdown("<div class='sidebar-section-title'>📅 TIME FILTER</div>", unsafe_allow_html=True)
    with st.sidebar.expander("Select Time Period", expanded=True):
        filter_level = st.radio(
            "Filter Level",
            ["Year", "Month", "Day"],
            index=["Year", "Month", "Day"].index(st.session_state.filter_level),
            horizontal=True,
            key="filter_level_radio"
        )
        st.session_state.filter_level = filter_level
        
        all_years = [2026]
        selected_year = st.selectbox("Select Year", all_years, index=0, key="year_select")
        st.session_state.selected_year = selected_year
        
        if filter_level in ["Month", "Day"]:
            months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            selected_month = st.selectbox("Select Month", range(1, 13), format_func=lambda x: months[x-1], index=st.session_state.selected_month - 1, key="month_select")
            st.session_state.selected_month = selected_month
            
            if filter_level == "Day":
                max_day = 31 if selected_month not in [4, 6, 9, 11] else 30
                if selected_month == 2:
                    max_day = 29
                selected_day = st.selectbox("Select Day", range(1, max_day + 1), index=min(st.session_state.selected_day - 1, max_day - 1), key="day_select")
                st.session_state.selected_day = selected_day
    
    # SECTION 7: NAVIGATION (subset)
    st.sidebar.markdown("<div class='sidebar-section-title'>🧭 NAVIGATION</div>", unsafe_allow_html=True)
    if st.session_state.user_role == "Admin":
        menu_options = ["Dashboard", "Monthly Team Evaluation", "Records", "Water Quality Data"]
    else:
        menu_options = ["Dashboard", "Records", "Water Quality Data"]
    selected_menu = st.sidebar.radio("Select Page", menu_options, key="menu_radio", label_visibility="collapsed")
    st.session_state.menu_option = selected_menu

    # ================= GET DATA FOR DISPLAY =================
    all_data = get_records(selected_branch)
    today = date.today()
    
    filtered_data = all_data.copy() if not all_data.empty else pd.DataFrame()
    
    if not filtered_data.empty:
        filtered_data["Date"] = pd.to_datetime(filtered_data["Date"])
        filtered_data = filtered_data[filtered_data["Date"].dt.year == selected_year]
        if st.session_state.filter_level in ["Month", "Day"]:
            filtered_data = filtered_data[filtered_data["Date"].dt.month == selected_month]
        if st.session_state.filter_level == "Day":
            filtered_data = filtered_data[filtered_data["Date"].dt.day == selected_day]
    
    # Monthly aggregation for Year view
    if st.session_state.filter_level == "Year" and not filtered_data.empty and len(filtered_data) > 0:
        try:
            monthly_data = filtered_data.copy()
            monthly_data['Month'] = monthly_data['Date'].dt.to_period('M')
            monthly_agg = monthly_data.groupby('Month').agg({
                'inflow_m3': 'sum',
                'outflow_m3': 'sum',
                'revenue_water_m3': 'sum',
                'total_production_cost': 'sum',
                'water_revenue': 'sum',
                'efficiency': 'mean',
                'Date': 'first'
            }).reset_index()
            monthly_agg['Date'] = monthly_agg['Month'].dt.start_time
            plot_data = monthly_agg
            is_monthly_agg = True
        except Exception as e:
            plot_data = filtered_data
            is_monthly_agg = False
    else:
        plot_data = filtered_data
        is_monthly_agg = False
    
    # ================= GET DATE RANGE =================
    date_display, date_sub = get_date_range_display(filtered_data)
    
    if not filtered_data.empty:
        day_perf = calculate_shift_summary_for_display(filtered_data, "Day", active_global_cost)
        night_perf = calculate_shift_summary_for_display(filtered_data, "Night", active_global_cost)
        display_df = filtered_data.copy()
        display_df.index = np.arange(1, len(display_df) + 1)
        
        total_raw = day_perf["total_raw_water"] + night_perf["total_raw_water"]
        total_treated = day_perf["total_treated"] + night_perf["total_treated"]
        total_expense = day_perf["total_expense"] + night_perf["total_expense"]
        total_revenue = day_perf["total_revenue"] + night_perf["total_revenue"]
        total_revenue_water = day_perf["revenue_water"] + night_perf["revenue_water"]
        total_backwashing = day_perf["backwashing"] + night_perf["backwashing"]
        
        # Capital and Prediction
        day_capital = day_perf["total_expense"] / day_perf["revenue_water"] if day_perf["revenue_water"] > 0 else 0
        night_capital = night_perf["total_expense"] / night_perf["revenue_water"] if night_perf["revenue_water"] > 0 else 0
        day_prediction = day_capital * 1.10 if day_capital > 0 else 0
        night_prediction = night_capital * 1.10 if night_capital > 0 else 0
        capital_val = total_expense / total_revenue_water if total_revenue_water > 0 else 0
        prediction_val = capital_val * 1.10 if capital_val > 0 else 0
        daily_avg = (day_prediction + night_prediction) / 2 if (day_prediction > 0 or night_prediction > 0) else 0
        
        # Collect all daily averages for the period
        all_daily_averages = []
        if not filtered_data.empty:
            shift_col = 'Shift' if 'Shift' in filtered_data.columns else 'shift' if 'shift' in filtered_data.columns else None
            
            for single_date in filtered_data['Date'].dt.date.unique():
                day_data = filtered_data[filtered_data['Date'].dt.date == single_date]
                
                if shift_col:
                    day_shift = day_data[day_data[shift_col] == 'Day']
                    night_shift = day_data[day_data[shift_col] == 'Night']
                else:
                    day_shift = day_data
                    night_shift = pd.DataFrame()
                
                day_shift_expense = day_shift['total_production_cost'].sum() if 'total_production_cost' in day_shift.columns else 0
                day_shift_rev_water = day_shift['revenue_water_m3'].sum() if 'revenue_water_m3' in day_shift.columns else 0
                day_cap = day_shift_expense / day_shift_rev_water if day_shift_rev_water > 0 else 0
                day_pred = day_cap * 1.10 if day_cap > 0 else 0
                
                night_shift_expense = night_shift['total_production_cost'].sum() if 'total_production_cost' in night_shift.columns else 0
                night_shift_rev_water = night_shift['revenue_water_m3'].sum() if 'revenue_water_m3' in night_shift.columns else 0
                night_cap = night_shift_expense / night_shift_rev_water if night_shift_rev_water > 0 else 0
                night_pred = night_cap * 1.10 if night_cap > 0 else 0
                
                if day_pred > 0 and night_pred > 0:
                    daily_avg_for_day = (day_pred + night_pred) / 2
                elif day_pred > 0:
                    daily_avg_for_day = day_pred
                elif night_pred > 0:
                    daily_avg_for_day = night_pred
                else:
                    daily_avg_for_day = 0
                
                if daily_avg_for_day > 0:
                    all_daily_averages.append(daily_avg_for_day)
        
        if st.session_state.filter_level == "Day":
            if all_daily_averages:
                avg_all = sum(all_daily_averages) / len(all_daily_averages)
            else:
                avg_all = daily_avg
            avg_label = "DAILY AVERAGE"
        elif st.session_state.filter_level == "Month":
            if all_daily_averages:
                avg_all = sum(all_daily_averages) / len(all_daily_averages)
            else:
                avg_all = 0
            avg_label = "MONTHLY AVERAGE"
        else:
            if all_daily_averages:
                avg_all = sum(all_daily_averages) / len(all_daily_averages)
            else:
                avg_all = 0
            avg_label = "YEARLY AVERAGE"
        
        latest_eff = (total_treated / total_raw * 100) if total_raw > 0 else 0
        revenue_eff = (total_revenue_water / total_treated * 100) if total_treated > 0 else 0
        latest_cost = total_revenue
        
        display_info = date_display
        has_data = True
    else:
        day_perf = {"total_raw_water": 0, "total_treated": 0, "total_expense": 0, 
                    "total_revenue": 0, "efficiency": 0, "quality_efficiency": 0,
                    "revenue_water": 0, "backwashing": 0, "record_count": 0}
        night_perf = day_perf.copy()
        display_df = pd.DataFrame()
        capital_val = prediction_val = avg_all = latest_eff = latest_cost = 0
        total_raw = total_treated = total_expense = total_revenue = 0
        total_revenue_water = total_backwashing = 0
        revenue_eff = 0
        display_info = "No Data Available"
        has_data = False
        plot_data = pd.DataFrame()
        is_monthly_agg = False
        day_prediction = 0
        night_prediction = 0
        avg_label = "AVERAGE"
        date_sub = ""

    daily_total_raw = day_perf['total_raw_water'] + night_perf['total_raw_water']
    daily_total_treated = day_perf['total_treated'] + night_perf['total_treated']
    daily_total_expense = day_perf['total_expense'] + night_perf['total_expense']
    daily_total_revenue = day_perf['total_revenue'] + night_perf['total_revenue']
    daily_revenue_water = day_perf['revenue_water'] + night_perf['revenue_water']
    daily_backwashing = day_perf['backwashing'] + night_perf['backwashing']
    daily_avg_efficiency = (day_perf['efficiency'] + night_perf['efficiency']) / 2 if (day_perf['efficiency'] > 0 or night_perf['efficiency'] > 0) else 0
    revenue_efficiency = (daily_revenue_water / daily_total_treated * 100) if daily_total_treated > 0 else 0

    # Display with Date Range
    st.markdown(f"""
    <div class='branch-header'>
        <span>NZOVE Water Treatment Plant</span><br>
        <div style='font-size: 14px; color: #1e3c72; margin-top: 5px;'>
            📅 {display_info}
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ================= DASHBOARD =================
    if st.session_state.menu_option == "Dashboard":
        left_col, right_col = st.columns([0.45, 0.55])

        with left_col:
            st.markdown("### Key Performance Indicators")
            
            # Row 1: BASE PRICE, CAPITAL, PREDICTION
            r1c1, r1c2, r1c3 = st.columns(3)
            with r1c1:
                st.markdown(f"""
                <div class='card'>
                    <div class='card-label'>BASE PRICE</div>
                    <div class='card-val'>{active_global_cost:,.0f}</div>
                    <div class='card-unit'>RWF / m³</div>
                </div>
                """, unsafe_allow_html=True)
            with r1c2:
                st.markdown(f"""
                <div class='card'>
                    <div class='card-label'>CAPITAL / m³</div>
                    <div class='card-val'>{capital_val:,.1f}</div>
                    <div class='card-unit'>RWF / m³</div>
                </div>
                """, unsafe_allow_html=True)
            with r1c3:
                st.markdown(f"""
                <div class='card'>
                    <div class='card-label'>PREDICTION</div>
                    <div class='card-val'>{prediction_val:,.1f}</div>
                    <div class='card-unit'>RWF / m³</div>
                </div>
                """, unsafe_allow_html=True)
            
            # Row 2: AVERAGE, EFFICIENCY, TOTAL REVENUE
            r2c1, r2c2, r2c3 = st.columns(3)
            with r2c1:
                avg_display = f"{avg_all:,.1f}" if avg_all > 0 else "0.0"
                st.markdown(f"""
                <div class='card' style='border-color: #8e44ad; background: linear-gradient(135deg, rgba(255,255,255,0.98), rgba(240,248,255,0.98));'>
                    <div class='card-label' style='color: #8e44ad;'>{avg_label}</div>
                    <div class='card-val' style='background: linear-gradient(135deg, #8e44ad, #9b59b6); -webkit-background-clip: text; -webkit-text-fill-color: transparent;'>{avg_display}</div>
                    <div class='card-unit'>RWF / m³</div>
                </div>
                """, unsafe_allow_html=True)
            with r2c2:
                eff_color = "green" if latest_eff >= 75 else "red" if latest_eff < 60 else "yellow"
                st.markdown(f"""
                <div class='card'>
                    <div class='card-label'>EFFICIENCY</div>
                    <div class='card-val'>{latest_eff:.1f}%</div>
                    <div class='card-unit'><span class='status-dot {eff_color}'></span> Recovery Rate</div>
                </div>
                """, unsafe_allow_html=True)
            with r2c3:
                st.markdown(f"""
                <div class='card'>
                    <div class='card-label'>TOTAL REVENUE</div>
                    <div class='card-val-large'>{format_number(latest_cost)}</div>
                    <div class='card-unit'>RWF</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
            st.markdown("### Water Production Metrics")
            
            r3c1, r3c2, r3c3, r3c4 = st.columns(4)
            with r3c1:
                st.markdown(f"""
                <div class='card' style='min-height: 80px;'>
                    <div class='card-label'>TOTAL TREATED</div>
                    <div class='card-val'>{format_number(daily_total_treated)}</div>
                    <div class='card-unit'>m³</div>
                </div>
                """, unsafe_allow_html=True)
            with r3c2:
                st.markdown(f"""
                <div class='card' style='min-height: 80px;'>
                    <div class='card-label'>REVENUE WATER</div>
                    <div class='card-val'>{format_number(daily_revenue_water)}</div>
                    <div class='card-unit'>m³ ({revenue_efficiency:.1f}%)</div>
                </div>
                """, unsafe_allow_html=True)
            with r3c3:
                st.markdown(f"""
                <div class='card' style='min-height: 80px;'>
                    <div class='card-label'>BACKWASHING</div>
                    <div class='card-val'>{format_number(daily_backwashing)}</div>
                    <div class='card-unit'>m³</div>
                </div>
                """, unsafe_allow_html=True)
            with r3c4:
                non_rev_val = daily_total_treated * (non_revenue_pct / 100)
                st.markdown(f"""
                <div class='card' style='min-height: 80px;'>
                    <div class='card-label'>NON-REVENUE</div>
                    <div class='card-val'>{format_number(non_rev_val)}</div>
                    <div class='card-unit'>{non_revenue_pct:.1f}% of treated</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
            
            # Shift Performance
            st.markdown("### Shift Performance Comparison")
            col_sum1, col_sum2, col_sum3, col_sum4, col_sum5 = st.columns(5)
            with col_sum1:
                st.markdown(f"""
                <div class='shift-box' style='border-color:#1e3c72;'>
                    <div style='font-size:9px;color:#1e3c72;'>TOTAL RAW</div>
                    <div style='font-size:13px;font-weight:800;color:#1e3c72;'>{format_number(daily_total_raw)}</div>
                    <div style='font-size:8px;color:#64748b;'>m³</div>
                </div>
                """, unsafe_allow_html=True)
            with col_sum2:
                st.markdown(f"""
                <div class='shift-box' style='border-color:#27ae60;'>
                    <div style='font-size:9px;color:#1e3c72;'>TREATED</div>
                    <div style='font-size:13px;font-weight:800;color:#27ae60;'>{format_number(daily_total_treated)}</div>
                    <div style='font-size:8px;color:#64748b;'>m³</div>
                </div>
                """, unsafe_allow_html=True)
            with col_sum3:
                st.markdown(f"""
                <div class='shift-box' style='border-color:#e74c3c;'>
                    <div style='font-size:9px;color:#1e3c72;'>EXPENSE</div>
                    <div style='font-size:13px;font-weight:800;color:#e74c3c;'>{format_number(daily_total_expense)}</div>
                    <div style='font-size:8px;color:#64748b;'>RWF</div>
                </div>
                """, unsafe_allow_html=True)
            with col_sum4:
                st.markdown(f"""
                <div class='shift-box' style='border-color:#f39c12;'>
                    <div style='font-size:9px;color:#1e3c72;'>REVENUE</div>
                    <div style='font-size:13px;font-weight:800;color:#f39c12;'>{format_number(daily_total_revenue)}</div>
                    <div style='font-size:8px;color:#64748b;'>RWF</div>
                </div>
                """, unsafe_allow_html=True)
            with col_sum5:
                st.markdown(f"""
                <div class='shift-box' style='border-color:#8e44ad;'>
                    <div style='font-size:9px;color:#1e3c72;'>AVG EFF</div>
                    <div style='font-size:13px;font-weight:800;color:#8e44ad;'>{daily_avg_efficiency:.1f}%</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # DAY SHIFT
            with st.expander("DAY SHIFT", expanded=True):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown(f"""
                    <div class='shift-box' style='border-color:#1e3c72;'>
                        <div style='font-size:10px;color:#1e3c72;'>RAW</div>
                        <div style='font-size:15px;font-weight:800;color:#1e3c72;'>{format_number(day_perf['total_raw_water'])}</div>
                        <div style='font-size:9px;color:#64748b;'>m³</div>
                    </div>
                    <div class='shift-box' style='border-color:#27ae60;'>
                        <div style='font-size:10px;color:#1e3c72;'>TREATED</div>
                        <div style='font-size:15px;font-weight:800;color:#27ae60;'>{format_number(day_perf['total_treated'])}</div>
                        <div style='font-size:9px;color:#64748b;'>m³</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col2:
                    st.markdown(f"""
                    <div class='shift-box' style='border-color:#e74c3c;'>
                        <div style='font-size:10px;color:#1e3c72;'>EXPENSE</div>
                        <div style='font-size:15px;font-weight:800;color:#e74c3c;'>{format_number(day_perf['total_expense'])}</div>
                        <div style='font-size:9px;color:#64748b;'>RWF</div>
                    </div>
                    <div class='shift-box' style='border-color:#f39c12;'>
                        <div style='font-size:10px;color:#1e3c72;'>REVENUE</div>
                        <div style='font-size:15px;font-weight:800;color:#f39c12;'>{format_number(day_perf['total_revenue'])}</div>
                        <div style='font-size:9px;color:#64748b;'>RWF</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col3:
                    st.markdown(f"""
                    <div class='shift-box' style='border-color:#1e3c72;'>
                        <div style='font-size:10px;color:#1e3c72;'>EFF</div>
                        <div style='font-size:15px;font-weight:800;color:#1e3c72;'>{day_perf['efficiency']:.1f}%</div>
                    </div>
                    <div class='shift-box' style='border-color:#8e44ad;'>
                        <div style='font-size:10px;color:#1e3c72;'>QUALITY</div>
                        <div style='font-size:15px;font-weight:800;color:#8e44ad;'>{day_perf['quality_efficiency']:.1f}%</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            # NIGHT SHIFT
            with st.expander("NIGHT SHIFT", expanded=True):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown(f"""
                    <div class='shift-box' style='border-color:#1e3c72;'>
                        <div style='font-size:10px;color:#1e3c72;'>RAW</div>
                        <div style='font-size:15px;font-weight:800;color:#1e3c72;'>{format_number(night_perf['total_raw_water'])}</div>
                        <div style='font-size:9px;color:#64748b;'>m³</div>
                    </div>
                    <div class='shift-box' style='border-color:#27ae60;'>
                        <div style='font-size:10px;color:#1e3c72;'>TREATED</div>
                        <div style='font-size:15px;font-weight:800;color:#27ae60;'>{format_number(night_perf['total_treated'])}</div>
                        <div style='font-size:9px;color:#64748b;'>m³</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col2:
                    st.markdown(f"""
                    <div class='shift-box' style='border-color:#e74c3c;'>
                        <div style='font-size:10px;color:#1e3c72;'>EXPENSE</div>
                        <div style='font-size:15px;font-weight:800;color:#e74c3c;'>{format_number(night_perf['total_expense'])}</div>
                        <div style='font-size:9px;color:#64748b;'>RWF</div>
                    </div>
                    <div class='shift-box' style='border-color:#f39c12;'>
                        <div style='font-size:10px;color:#1e3c72;'>REVENUE</div>
                        <div style='font-size:15px;font-weight:800;color:#f39c12;'>{format_number(night_perf['total_revenue'])}</div>
                        <div style='font-size:9px;color:#64748b;'>RWF</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col3:
                    st.markdown(f"""
                    <div class='shift-box' style='border-color:#1e3c72;'>
                        <div style='font-size:10px;color:#1e3c72;'>EFF</div>
                        <div style='font-size:15px;font-weight:800;color:#1e3c72;'>{night_perf['efficiency']:.1f}%</div>
                    </div>
                    <div class='shift-box' style='border-color:#8e44ad;'>
                        <div style='font-size:10px;color:#1e3c72;'>QUALITY</div>
                        <div style='font-size:15px;font-weight:800;color:#8e44ad;'>{night_perf['quality_efficiency']:.1f}%</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            if day_perf['efficiency'] > night_perf['efficiency']:
                st.success(f"Day shift has {day_perf['efficiency'] - night_perf['efficiency']:.1f}% higher efficiency")
            elif night_perf['efficiency'] > day_perf['efficiency']:
                st.success(f"Night shift has {night_perf['efficiency'] - day_perf['efficiency']:.1f}% higher efficiency")
            else:
                st.info(f"Both shifts have similar efficiency")
            
            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
            
            # ================= DELETE RECORD - Today Only =================
            with st.expander("🗑️ DELETE ALL RECORDS (Today Only)", expanded=False):
                st.warning("⚠️ This will delete ALL records for today (Day + Night shifts)")
                st.info(f"📅 Date: **{today}**")
                
                today_records = get_records(selected_branch, year=today.year, month=today.month, day=today.day)
                
                if not today_records.empty:
                    st.markdown("**Records to be deleted:**")
                    for idx, record in today_records.iterrows():
                        shift_val = record.get('shift', record.get('Shift', 'N/A'))
                        st.markdown(f"- **{shift_val} shift** | Team: {record.get('team', 'N/A')} | Inflow: {record.get('inflow_m3', 0):.0f} m³ | Outflow: {record.get('outflow_m3', 0):.0f} m³")
                    
                    st.markdown("---")
                    st.warning("⚠️ Type 'DELETE' in the box below to confirm deletion of ALL records for today")
                    confirm_delete_today = st.text_input("Confirm with 'DELETE'", key="confirm_delete_today", type="password")
                    
                    if st.button("🗑️ DELETE ALL TODAY", use_container_width=True, key="delete_all_today_btn"):
                        if confirm_delete_today == "DELETE":
                            delete_record_for_date(today, selected_branch, "ALL")
                            st.success(f"✅ ALL records for {today} have been deleted!")
                            st.info("📝 All values are now zero. System is ready for new data.")
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error("❌ Please type 'DELETE' to confirm")
                else:
                    st.success("✅ No records for today. System is ready for new data.")
                    st.info("📝 All values are at zero. You can start adding new records.")
            
            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
            
            # ================= DATA ENTRY - ONLY ON DAY VIEW =================
            if st.session_state.filter_level == "Day":
                with st.expander("📝 ADD NEW RECORD (Today Only)", expanded=True):
                    if st.session_state.user_role == "Admin":
                        st.warning("⚠️ Admin cannot add records. Only Users can add records.")
                    else:
                        try:
                            today = date.today()
                            
                            record_date = st.date_input("Date", today, key="entry_date")
                            
                            if record_date != today:
                                st.error("❌ Only today's date is allowed! Please select today.")
                                st.stop()
                            
                            shift = st.selectbox("Shift", ["Day", "Night"], key="shift_select")
                            
                            if shift == "Night":
                                start_date_str = record_date.strftime("%Y/%m/%d")
                                end_date = record_date + timedelta(days=1)
                                end_date_str = end_date.strftime("%Y/%m/%d")
                                date_range_display = f"{start_date_str} - {end_date_str}"
                            else:
                                date_range_display = record_date.strftime("%Y/%m/%d")
                            
                            st.info(f"📅 Record Date: **{date_range_display}**")
                            st.caption(f"⏰ {shift} Shift")
                            
                            selected_team_type = st.selectbox("Select Team", TEAM_TYPES, key="record_team_select")
                            team_info = get_team_info(selected_branch, selected_team_type)
                            team_leader = team_info.get("team_leader", "Unknown") if team_info else "Unknown"
                            
                            st.markdown("#### Water Flow Tracking")
                            
                            meters_config = load_meters_config(selected_branch)
                            
                            # ===== RAW WATER METERS =====
                            st.markdown("**Raw Water Flowmeters**")
                            rw_inputs = []
                            total_raw_consumption = 0
                            
                            for idx, rw_m in enumerate(meters_config["raw_water"]):
                                prev_reading = get_latest_meter_reading(
                                    selected_branch, "raw_water", rw_m
                                )
                                prev_val = prev_reading["reading_value"] if prev_reading is not None else 0.0
                                
                                if prev_val > 0:
                                    prev_date = prev_reading["reading_date"] if prev_reading is not None else "N/A"
                                    st.markdown(f"""
                                    <div style='background-color: #000000; color: #FFFFFF; padding: 8px 12px; border-radius: 8px; margin-bottom: 8px; font-size: 14px;'>
                                        📊 {rw_m} - Previous: {prev_val:,.1f} m³ | Date: {prev_date}
                                    </div>
                                    """, unsafe_allow_html=True)
                                else:
                                    st.markdown(f"""
                                    <div style='background-color: #000000; color: #FFFFFF; padding: 8px 12px; border-radius: 8px; margin-bottom: 8px; font-size: 14px;'>
                                        📊 {rw_m} - No previous reading found
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                current_val = st.number_input(
                                    f"{rw_m} Current (m³)", 
                                    min_value=0.0, 
                                    value=0.0,
                                    step=10.0,
                                    key=f"rw_reading_{rw_m}_{idx}"
                                )
                                rw_inputs.append(current_val)
                                
                                if current_val > prev_val:
                                    consumption = current_val - prev_val
                                    total_raw_consumption += consumption
                                    if consumption > 0:
                                        st.caption(f"📈 Consumption: {consumption:,.1f} m³")
                            
                            rw_other = st.number_input("Other Raw Water (m³)", min_value=0.0, value=0.0, key="rw_other_dynamic")
                            total_raw_consumption += rw_other
                            calculated_inflow = total_raw_consumption
                            st.info(f"✅ Total Inflow: {calculated_inflow:,.0f} m³")
                            
                            # ===== TREATED WATER METERS =====
                            st.markdown("**Treated Water Flowmeters**")
                            tw_inputs = []
                            total_treated_consumption = 0
                            
                            for idx, tw_m in enumerate(meters_config["treated_water"]):
                                prev_reading = get_latest_meter_reading(
                                    selected_branch, "treated_water", tw_m
                                )
                                prev_val = prev_reading["reading_value"] if prev_reading is not None else 0.0
                                
                                if prev_val > 0:
                                    prev_date = prev_reading["reading_date"] if prev_reading is not None else "N/A"
                                    st.markdown(f"""
                                    <div style='background-color: #000000; color: #FFFFFF; padding: 8px 12px; border-radius: 8px; margin-bottom: 8px; font-size: 14px;'>
                                        📊 {tw_m} - Previous: {prev_val:,.1f} m³ | Date: {prev_date}
                                    </div>
                                    """, unsafe_allow_html=True)
                                else:
                                    st.markdown(f"""
                                    <div style='background-color: #000000; color: #FFFFFF; padding: 8px 12px; border-radius: 8px; margin-bottom: 8px; font-size: 14px;'>
                                        📊 {tw_m} - No previous reading found
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                current_val = st.number_input(
                                    f"{tw_m} Current (m³)", 
                                    min_value=0.0, 
                                    value=0.0,
                                    step=10.0,
                                    key=f"tw_reading_{tw_m}_{idx}"
                                )
                                tw_inputs.append(current_val)
                                
                                if current_val > prev_val:
                                    consumption = current_val - prev_val
                                    total_treated_consumption += consumption
                                    if consumption > 0:
                                        st.caption(f"📈 Consumption: {consumption:,.1f} m³")
                            
                            tw_other = st.number_input("Other Treated Water (m³)", min_value=0.0, value=0.0, key="tw_other_dynamic")
                            total_treated_consumption += tw_other
                            calculated_treated = total_treated_consumption
                            st.info(f"✅ Total Treated: {calculated_treated:,.0f} m³")
                            
                            # ===== BACKWASHING METERS =====
                            st.markdown("**Backwashing Water Meters**")
                            bw_inputs = []
                            total_backwash_consumption = 0
                            
                            for idx, bw_m in enumerate(meters_config["backwashing"]):
                                prev_reading = get_latest_meter_reading(
                                    selected_branch, "backwashing", bw_m
                                )
                                prev_val = prev_reading["reading_value"] if prev_reading is not None else 0.0
                                
                                if prev_val > 0:
                                    prev_date = prev_reading["reading_date"] if prev_reading is not None else "N/A"
                                    st.markdown(f"""
                                    <div style='background-color: #000000; color: #FFFFFF; padding: 8px 12px; border-radius: 8px; margin-bottom: 8px; font-size: 14px;'>
                                        📊 {bw_m} - Previous: {prev_val:,.1f} m³ | Date: {prev_date}
                                    </div>
                                    """, unsafe_allow_html=True)
                                else:
                                    st.markdown(f"""
                                    <div style='background-color: #000000; color: #FFFFFF; padding: 8px 12px; border-radius: 8px; margin-bottom: 8px; font-size: 14px;'>
                                        📊 {bw_m} - No previous reading found
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                current_val = st.number_input(
                                    f"{bw_m} Current (m³)", 
                                    min_value=0.0, 
                                    value=0.0,
                                    step=10.0,
                                    key=f"bw_reading_{bw_m}_{idx}"
                                )
                                bw_inputs.append(current_val)
                                
                                if current_val > prev_val:
                                    consumption = current_val - prev_val
                                    total_backwash_consumption += consumption
                                    if consumption > 0:
                                        st.caption(f"📈 Consumption: {consumption:,.1f} m³")
                            
                            backwashing_val = total_backwash_consumption
                            st.info(f"✅ Total Backwashing: {backwashing_val:,.0f} m³")
                            
                            # ===== REVENUE WATER =====
                            non_rev_pct = float(current_thresholds.get("Non_Revenue_Pct", 10.0))
                            non_rev_val = calculated_treated * (non_rev_pct / 100)
                            calculated_revenue_water = calculated_treated - backwashing_val - non_rev_val
                            st.info(f"✅ Revenue Water: {calculated_revenue_water:,.0f} m³")
                            
                            # ===== PRODUCTION COSTS =====
                            st.markdown("#### Production Costs")
                            
                            st.markdown("**Chemicals**")
                            col_c1, col_c2, col_c3 = st.columns(3)
                            with col_c1:
                                chem_lime = st.number_input("Lime (kg)", min_value=0.0, value=0.0, key="c_lime")
                            with col_c2:
                                chem_fl4440 = st.number_input("FL4440 (L)", min_value=0.0, value=0.0, key="c_fl")
                            with col_c3:
                                chem_nacl = st.number_input("NaCl (kg)", min_value=0.0, value=0.0, key="c_nacl")
                            
                            price_lime = float(current_thresholds.get("Price_Lime", 450.0))
                            price_fl = float(current_thresholds.get("Price_FL4440", 2500.0))
                            price_nacl = float(current_thresholds.get("Price_NaCl", 300.0))
                            price_elec = float(current_thresholds.get("Price_Electricity", 148.0))
                            
                            subtotal_chem = (chem_lime * price_lime) + (chem_fl4440 * price_fl) + (chem_nacl * price_nacl)
                            st.markdown(f"<div class='subtotal-container'>Chemicals: {subtotal_chem:,.0f} RWF</div>", unsafe_allow_html=True)
                            
                            # ===== ELECTRICITY METERS =====
                            st.markdown("**Energy**")
                            elec_inputs = []
                            total_elec_consumption = 0
                            
                            for idx, el_m in enumerate(meters_config["electricity"]):
                                prev_reading = get_latest_meter_reading(
                                    selected_branch, "electricity", el_m
                                )
                                prev_val = prev_reading["reading_value"] if prev_reading is not None else 0.0
                                
                                if prev_val > 0:
                                    prev_date = prev_reading["reading_date"] if prev_reading is not None else "N/A"
                                    st.markdown(f"""
                                    <div style='background-color: #000000; color: #FFFFFF; padding: 8px 12px; border-radius: 8px; margin-bottom: 8px; font-size: 14px;'>
                                        📊 {el_m} - Previous: {prev_val:,.1f} kWh | Date: {prev_date}
                                    </div>
                                    """, unsafe_allow_html=True)
                                else:
                                    st.markdown(f"""
                                    <div style='background-color: #000000; color: #FFFFFF; padding: 8px 12px; border-radius: 8px; margin-bottom: 8px; font-size: 14px;'>
                                        📊 {el_m} - No previous reading found
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                current_val = st.number_input(
                                    f"{el_m} Current (kWh)", 
                                    min_value=0.0, 
                                    value=0.0,
                                    step=10.0,
                                    key=f"el_reading_{el_m}_{idx}"
                                )
                                elec_inputs.append(current_val)
                                
                                if current_val > prev_val:
                                    consumption = current_val - prev_val
                                    total_elec_consumption += consumption
                                    if consumption > 0:
                                        st.caption(f"📈 Consumption: {consumption:,.1f} kWh")
                            
                            elec_other = st.number_input("Other (kWh)", min_value=0.0, value=0.0, key="el_other_dynamic")
                            total_elec_consumption += elec_other
                            subtotal_energy = total_elec_consumption * price_elec
                            st.markdown(f"<div class='subtotal-container'>Energy ({total_elec_consumption:.0f} kWh): {subtotal_energy:,.0f} RWF</div>", unsafe_allow_html=True)
                            
                            # ===== OPERATIONAL COSTS =====
                            st.markdown("**Operational**")
                            col_o1, col_o2, col_o3 = st.columns(3)
                            with col_o1:
                                labor_cost = st.number_input("Labor (RWF)", min_value=0.0, value=0.0, key="c_labor")
                            with col_o2:
                                maint_cost = st.number_input("Maintenance (RWF)", min_value=0.0, value=0.0, key="c_maint")
                            with col_o3:
                                other_costs = st.number_input("Other (RWF)", min_value=0.0, value=0.0, key="c_other")
                            
                            calculated_production_cost = subtotal_chem + subtotal_energy + labor_cost + maint_cost + other_costs
                            st.success(f"✅ Total Cost: {calculated_production_cost:,.0f} RWF")
                            
                            # ===== TEAM COMMENT =====
                            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
                            st.markdown("#### Team Comment")
                            team_comment = st.text_area(
                                "Add any observations, challenges, or reasons for performance",
                                placeholder="e.g., Low water pressure, chemical shortage, equipment issue...",
                                height=80,
                                key=f"team_comment_entry_{selected_team_type}"
                            )
                            
                            # ===== WATER QUALITY =====
                            st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
                            st.markdown("#### Water Quality Data")
                            
                            with st.container():
                                st.markdown("<div class='water-quality-card'>", unsafe_allow_html=True)
                                
                                col_wq1, col_wq2 = st.columns(2)
                                with col_wq1:
                                    wq_date = st.date_input("Date", record_date, key="wq_date_entry")
                                    current_hour = datetime.now().hour
                                    allowed_hours = [current_hour]
                                    st.info(f"⏰ Only current hour ({current_hour:02d}:00) is allowed")
                                    wq_hour = st.selectbox(
                                        "Hour", 
                                        allowed_hours,
                                        format_func=lambda x: f"{x:02d}:00",
                                        key="wq_hour_entry"
                                    )
                                    flow_m3 = st.number_input("Flow (m³/h)", min_value=0.0, value=0.0, step=10.0, key="wq_flow_entry")
                                    raw_ntu = st.number_input("Raw Turbidity (NTU)", min_value=0.0, value=0.0, step=5.0, key="wq_ntu_entry")
                                with col_wq2:
                                    raw_ph = st.number_input("pH", min_value=0.0, max_value=14.0, value=7.0, step=0.1, key="wq_ph_entry")
                                    stock_pct = st.number_input("Stock Concentration (%)", min_value=0.0, value=0.0, step=0.1, key="wq_stock_entry")
                                    raw_temp = st.number_input("Temperature (°C)", min_value=0.0, value=20.0, step=0.5, key="wq_temp_entry")
                                    raw_chlorine = st.number_input("Chlorine (mg/L)", min_value=0.0, value=0.0, step=0.1, key="wq_chlorine_entry")
                                    raw_nitrate = st.number_input("Nitrate (mg/L)", min_value=0.0, value=0.0, step=1.0, key="wq_nitrate_entry")
                                
                                if st.button("Calculate & Save Water Quality", use_container_width=True, key="wq_save_btn"):
                                    if flow_m3 <= 0 or raw_ntu <= 0:
                                        st.error("⚠️ Flow and Raw Turbidity must be greater than 0!")
                                    else:
                                        jar_result = automatic_jar_test(flow_m3, raw_ntu, stock_pct)
                                        
                                        st.markdown("#### Results")
                                        col_r1, col_r2, col_r3 = st.columns(3)
                                        with col_r1:
                                            st.metric("Dose", f"{jar_result['dose_mgL']} mg/L")
                                            st.metric("Outlet NTU", f"{jar_result['outlet_ntu']} NTU")
                                        with col_r2:
                                            st.metric("Efficiency", f"{jar_result['efficiency']}%")
                                            st.metric("Pump", f"{jar_result['pump_Lh']} L/h")
                                        with col_r3:
                                            st.metric("Pump (L/min)", f"{jar_result['pump_Lmin']:.1f} L/min")
                                        
                                        datetime_str = f"{wq_date} {wq_hour:02d}:00:00"
                                        save_water_quality_record(
                                            datetime_str, selected_branch, flow_m3, raw_ntu, stock_pct,
                                            jar_result['dose_mgL'], jar_result['outlet_ntu'], jar_result['efficiency'],
                                            jar_result['pump_Lh'], jar_result['pump_Lmin'],
                                            raw_ph, raw_temp, raw_chlorine, raw_nitrate, st.session_state.current_user
                                        )
                                        st.success(f"✅ Water quality saved for {wq_date} at {wq_hour:02d}:00")
                                        st.balloons()
                                
                                st.markdown("</div>", unsafe_allow_html=True)
                            
                            # ===== SAVE PRODUCTION RECORD =====
                            if st.button("Save Production Record", use_container_width=True, key="save_record_btn"):
                                existing_data = get_records(selected_branch, year=record_date.year, month=record_date.month, day=record_date.day)
                                shift_col = 'shift' if 'shift' in existing_data.columns else 'Shift' if 'Shift' in existing_data.columns else None
                                duplicate = False
                                if shift_col and not existing_data.empty:
                                    duplicate = (existing_data[shift_col] == shift).any()
                                
                                if duplicate:
                                    st.error(f"❌ Record already exists for {record_date} ({shift})")
                                elif calculated_inflow <= 0 or calculated_treated <= 0:
                                    st.error("❌ Inflow and Treated must be > 0")
                                elif calculated_revenue_water < 0:
                                    st.error("❌ Revenue Water cannot be negative")
                                else:
                                    calc_revenue = calculated_revenue_water * active_global_cost
                                    efficiency = (calculated_treated / calculated_inflow * 100)
                                    
                                    record_data = {
                                        "date": str(record_date),
                                        "branch": selected_branch,
                                        "shift": shift,
                                        "date_range": date_range_display,
                                        "team": selected_team_type,
                                        "inflow_m3": calculated_inflow,
                                        "outflow_m3": calculated_treated,
                                        "backwashing_m3": backwashing_val,
                                        "non_revenue_pct": non_rev_pct,
                                        "revenue_water_m3": calculated_revenue_water,
                                        "loss_m3": (calculated_inflow - calculated_treated),
                                        "total_production_cost": calculated_production_cost,
                                        "water_revenue": calc_revenue,
                                        "net_profit": (calc_revenue - calculated_production_cost),
                                        "recorded_by": st.session_state.current_user,
                                        "efficiency": efficiency
                                    }
                                    
                                    save_record(record_data)
                                    
                                    save_team_performance(
                                        record_date, selected_branch, selected_team_type, team_leader, shift, 
                                        calculated_inflow, calculated_treated, 
                                        backwashing_val, non_rev_pct, calculated_revenue_water,
                                        calculated_production_cost, calc_revenue, 
                                        efficiency, 100, st.session_state.current_user,
                                        team_comment
                                    )
                                    
                                    meters_config = load_meters_config(selected_branch)
                                    
                                    for idx, rw_m in enumerate(meters_config["raw_water"]):
                                        val = rw_inputs[idx]
                                        if val > 0:
                                            save_meter_reading(
                                                str(record_date), selected_branch, "raw_water", rw_m, val, 
                                                st.session_state.current_user
                                            )
                                    
                                    for idx, tw_m in enumerate(meters_config["treated_water"]):
                                        val = tw_inputs[idx]
                                        if val > 0:
                                            save_meter_reading(
                                                str(record_date), selected_branch, "treated_water", tw_m, val, 
                                                st.session_state.current_user
                                            )
                                    
                                    for idx, bw_m in enumerate(meters_config["backwashing"]):
                                        val = bw_inputs[idx]
                                        if val > 0:
                                            save_meter_reading(
                                                str(record_date), selected_branch, "backwashing", bw_m, val, 
                                                st.session_state.current_user
                                            )
                                    
                                    for idx, el_m in enumerate(meters_config["electricity"]):
                                        val = elec_inputs[idx]
                                        if val > 0:
                                            save_meter_reading(
                                                str(record_date), selected_branch, "electricity", el_m, val, 
                                                st.session_state.current_user
                                            )
                                    
                                    st.success(f"✅ Record saved for {date_range_display} (Team {selected_team_type})!")
                                    if team_comment:
                                        st.info(f"📝 Comment saved: {team_comment}")
                                    st.cache_data.clear()
                                    st.balloons()
                        except Exception as e:
                            st.error(f"Error: {str(e)}")
            else:
                st.info("📝 Data entry is only available when viewing by DAY. Please select 'Day' in the filter.")

        # ================= RIGHT COLUMN - GRAPHS =================
        with right_col:
            if has_data and not plot_data.empty:
                plot_df_sorted = plot_data.sort_values("Date")
                plt.style.use("default")
                
                if is_monthly_agg:
                    dates = plot_df_sorted['Date'].dt.strftime("%b %Y")
                else:
                    dates = plot_df_sorted["Date"].dt.strftime("%m-%d")
                
                is_day_filter = st.session_state.filter_level == "Day"
                
                shift_col = 'Shift' if 'Shift' in filtered_data.columns else 'shift' if 'shift' in filtered_data.columns else None
                
                # ===== WATER FLOW GRAPH (Day, Night, Daily) =====
                if is_day_filter and shift_col:
                    day_data = filtered_data[filtered_data[shift_col] == "Day"]
                    night_data = filtered_data[filtered_data[shift_col] == "Night"]
                    
                    day_inflow = day_data['inflow_m3'].sum() if not day_data.empty else 0
                    day_outflow = day_data['outflow_m3'].sum() if not day_data.empty else 0
                    day_revenue = day_data['revenue_water_m3'].sum() if not day_data.empty else 0
                    
                    night_inflow = night_data['inflow_m3'].sum() if not night_data.empty else 0
                    night_outflow = night_data['outflow_m3'].sum() if not night_data.empty else 0
                    night_revenue = night_data['revenue_water_m3'].sum() if not night_data.empty else 0
                    
                    daily_inflow = day_inflow + night_inflow
                    daily_outflow = day_outflow + night_outflow
                    daily_revenue = day_revenue + night_revenue
                    
                    st.markdown("<div class='graph-container'>", unsafe_allow_html=True)
                    st.markdown("<div class='graph-title'>WATER FLOW BY SHIFT</div>", unsafe_allow_html=True)
                    
                    fig_shift, ax_shift = plt.subplots(figsize=(10, 4))
                    ax_shift.set_facecolor('#f0f8ff')
                    fig_shift.patch.set_facecolor('#87CEEB')
                    
                    x_pos = np.arange(3)
                    width = 0.25
                    
                    ax_shift.bar(x_pos - width, [day_inflow, night_inflow, daily_inflow], width, 
                                label='Raw Water', color='#1e3c72', alpha=0.8)
                    ax_shift.bar(x_pos, [day_outflow, night_outflow, daily_outflow], width, 
                                label='Treated', color='#e67e22', alpha=0.8)
                    ax_shift.bar(x_pos + width, [day_revenue, night_revenue, daily_revenue], width, 
                                label='Revenue Water', color='#27ae60', alpha=0.8)
                    
                    ax_shift.set_xlabel('Shift', fontsize=12)
                    ax_shift.set_ylabel('Volume (m³)', fontsize=12)
                    ax_shift.set_title(f'Nzove - Water Flow by Shift', 
                                      fontsize=14, weight='bold', color='#1e3c72')
                    ax_shift.set_xticks(x_pos)
                    ax_shift.set_xticklabels(['Day Shift', 'Night Shift', 'Daily Total'])
                    ax_shift.legend(loc='upper right')
                    ax_shift.grid(True, linestyle='--', alpha=0.3)
                    
                    plt.tight_layout()
                    st.pyplot(fig_shift)
                    
                    st.markdown(f"""
                    <div class='graph-comment'>
                    <b>Daily Summary - {date_display}</b><br>
                    • <b>Day Shift:</b> Raw: {day_inflow:,.0f} m³ | Treated: {day_outflow:,.0f} m³ | Revenue: {day_revenue:,.0f} m³<br>
                    • <b>Night Shift:</b> Raw: {night_inflow:,.0f} m³ | Treated: {night_outflow:,.0f} m³ | Revenue: {night_revenue:,.0f} m³<br>
                    • <b>Daily Total:</b> Raw: {daily_inflow:,.0f} m³ | Treated: {daily_outflow:,.0f} m³ | Revenue: {daily_revenue:,.0f} m³
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown("</div>", unsafe_allow_html=True)
                
                if not is_day_filter:
                    st.markdown("<div class='graph-container'>", unsafe_allow_html=True)
                    st.markdown("<div class='graph-title'>WATER FLOW</div>", unsafe_allow_html=True)
                    
                    fig1, ax1 = plt.subplots(figsize=(9, 3.5))
                    ax1.set_facecolor('#f0f8ff')
                    fig1.patch.set_facecolor('#87CEEB')
                    
                    flow_col = plot_df_sorted["inflow_m3"] if "inflow_m3" in plot_df_sorted.columns else pd.Series([0] * len(plot_df_sorted))
                    treated_col = plot_df_sorted["outflow_m3"] if "outflow_m3" in plot_df_sorted.columns else pd.Series([0] * len(plot_df_sorted))
                    revenue_col = plot_df_sorted["revenue_water_m3"] if "revenue_water_m3" in plot_df_sorted.columns else treated_col * 0.9
                    
                    ax1.plot(dates, flow_col, label="Raw Water", color="#1e3c72", marker='o', markersize=5, linewidth=2.5)
                    ax1.plot(dates, treated_col, label="Treated", color="#e67e22", marker='s', markersize=5, linewidth=2)
                    ax1.plot(dates, revenue_col, label="Revenue", color="#27ae60", marker='^', markersize=5, linewidth=2)
                    plt.xticks(rotation=45)
                    
                    ax1.set_title("Nzove - Water Flow", color="#1e3c72", fontsize=13, weight='bold')
                    ax1.set_xlabel("Date", fontsize=11, color="#1e3c72")
                    ax1.set_ylabel("Volume (m³)", fontsize=11, color="#1e3c72")
                    ax1.tick_params(colors="#1e3c72")
                    ax1.grid(True, linestyle='--', alpha=0.3)
                    ax1.legend(loc='upper right', fontsize=9)
                    plt.tight_layout()
                    st.pyplot(fig1)
                    
                    total_raw_val = flow_col.sum()
                    total_treated_val = treated_col.sum()
                    total_revenue_val = revenue_col.sum()
                    
                    st.markdown(f"""
                    <div class='graph-comment'>
                    <b>Water Flow Interpretation:</b><br>
                    • <b>Blue line (Raw Water):</b> Water entering the plant<br>
                    • <b>Orange line (Treated):</b> Water after treatment<br>
                    • <b>Green line (Revenue):</b> Water sold to customers<br>
                    <br>
                    <b>Total Values:</b><br>
                    • <b>Raw Water:</b> {format_number(total_raw_val)} m³<br>
                    • <b>Treated Water:</b> {format_number(total_treated_val)} m³<br>
                    • <b>Revenue Water:</b> {format_number(total_revenue_val)} m³
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown("</div>", unsafe_allow_html=True)
                
                # ===== FINANCIAL PERFORMANCE GRAPH (Day, Night, Daily) =====
                if "total_production_cost" in plot_df_sorted.columns and "water_revenue" in plot_df_sorted.columns:
                    st.markdown("<div class='graph-container'>", unsafe_allow_html=True)
                    st.markdown("<div class='graph-title'>FINANCIAL PERFORMANCE</div>", unsafe_allow_html=True)
                    
                    fig2, ax2 = plt.subplots(figsize=(9, 3.5))
                    ax2.set_facecolor('#f0f8ff')
                    fig2.patch.set_facecolor('#87CEEB')
                    
                    if is_day_filter and shift_col:
                        day_data = filtered_data[filtered_data[shift_col] == "Day"]
                        night_data = filtered_data[filtered_data[shift_col] == "Night"]
                        
                        day_cost = day_data['total_production_cost'].sum() if not day_data.empty else 0
                        day_revenue = day_data['water_revenue'].sum() if not day_data.empty else 0
                        night_cost = night_data['total_production_cost'].sum() if not night_data.empty else 0
                        night_revenue = night_data['water_revenue'].sum() if not night_data.empty else 0
                        daily_cost = day_cost + night_cost
                        daily_revenue = day_revenue + night_revenue
                        
                        x_pos = np.arange(3)
                        width = 0.25
                        
                        ax2.bar(x_pos - width/2, [day_cost, night_cost, daily_cost], width, 
                                label='Production Cost', color='#e74c3c', alpha=0.7)
                        ax2.bar(x_pos + width/2, [day_revenue, night_revenue, daily_revenue], width, 
                                label='Revenue', color='#27ae60', alpha=0.7)
                        
                        ax2.set_xticks(x_pos)
                        ax2.set_xticklabels(['Day Shift', 'Night Shift', 'Daily Total'])
                    else:
                        ax2.plot(dates, plot_df_sorted["total_production_cost"], label="Production Cost", color="#e74c3c", marker='^', markersize=5, linewidth=2)
                        ax2.plot(dates, plot_df_sorted["water_revenue"], label="Revenue", color="#27ae60", marker='s', markersize=5, linewidth=2)
                        plt.xticks(rotation=45)
                    
                    ax2.set_title("Nzove - Revenue vs Production Cost", color="#27ae60", fontsize=13, weight='bold')
                    ax2.set_xlabel("Date", fontsize=11, color="#1e3c72")
                    ax2.set_ylabel("Amount (RWF)", fontsize=11, color="#1e3c72")
                    ax2.tick_params(colors="#1e3c72")
                    ax2.grid(True, linestyle=':', alpha=0.3)
                    ax2.legend(loc='upper left', fontsize=10)
                    plt.tight_layout()
                    st.pyplot(fig2)
                    
                    total_rev = plot_df_sorted["water_revenue"].sum() if "water_revenue" in plot_df_sorted.columns else 0
                    total_cost = plot_df_sorted["total_production_cost"].sum() if "total_production_cost" in plot_df_sorted.columns else 0
                    profit = total_rev - total_cost
                    st.markdown(f"""
                    <div class='graph-comment'>
                    <b>Financial Interpretation:</b><br>
                    • <b>Red bars/line (Cost):</b> Total production cost<br>
                    • <b>Green bars/line (Revenue):</b> Total revenue<br>
                    • <b>Total Revenue:</b> {format_number(total_rev)} RWF<br>
                    • <b>Total Cost:</b> {format_number(total_cost)} RWF<br>
                    • <b>Net Profit:</b> {format_number(profit)} RWF
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown("</div>", unsafe_allow_html=True)
                
                # ===== OPERATIONAL EFFICIENCY GRAPH (Day, Night, Daily) =====
                if "efficiency" in plot_df_sorted.columns or "Efficiency_%" in plot_df_sorted.columns:
                    st.markdown("<div class='graph-container'>", unsafe_allow_html=True)
                    st.markdown("<div class='graph-title'>OPERATIONAL EFFICIENCY</div>", unsafe_allow_html=True)
                    
                    fig4, ax4 = plt.subplots(figsize=(9, 3.5))
                    ax4.set_facecolor('#f0f8ff')
                    fig4.patch.set_facecolor('#87CEEB')
                    
                    if "efficiency" in plot_df_sorted.columns:
                        eff_col = "efficiency"
                    else:
                        eff_col = "Efficiency_%"
                    
                    if is_day_filter and shift_col:
                        day_data = filtered_data[filtered_data[shift_col] == "Day"]
                        night_data = filtered_data[filtered_data[shift_col] == "Night"]
                        
                        day_eff = day_data[eff_col].mean() if not day_data.empty else 0
                        night_eff = night_data[eff_col].mean() if not night_data.empty else 0
                        daily_eff = (day_eff + night_eff) / 2 if (day_eff > 0 or night_eff > 0) else 0
                        
                        bars = ax4.bar(['Day Shift', 'Night Shift', 'Daily Total'], 
                                     [day_eff, night_eff, daily_eff], 
                                     color=["#1e3c72", "#2a5298", "#8e44ad"], alpha=0.7)
                        
                        for bar in bars:
                            if bar.get_height() >= 75:
                                bar.set_color('#27ae60')
                            elif bar.get_height() >= 60:
                                bar.set_color('#f39c12')
                            else:
                                bar.set_color('#e74c3c')
                        
                        ax4.axhline(y=75, color='#e74c3c', linestyle='--', alpha=0.7, linewidth=2, label="Target 75%")
                    else:
                        ax4.plot(dates, plot_df_sorted[eff_col], label="Efficiency", color="#1e3c72", marker='o', markersize=6, linewidth=2.5)
                        ax4.axhline(y=75, color='#e74c3c', linestyle='--', alpha=0.7, linewidth=2, label="Target 75%")
                        plt.xticks(rotation=45)
                    
                    ax4.set_title("Nzove - Operational Efficiency", color="#1e3c72", fontsize=13, weight='bold')
                    ax4.set_xlabel("Date", fontsize=11, color="#1e3c72")
                    ax4.set_ylabel("Efficiency (%)", fontsize=11, color="#1e3c72")
                    ax4.tick_params(colors="#1e3c72")
                    ax4.grid(True, linestyle='--', alpha=0.3)
                    ax4.legend(loc='upper left', fontsize=10)
                    plt.tight_layout()
                    st.pyplot(fig4)
                    
                    avg_eff = plot_df_sorted[eff_col].mean() if eff_col in plot_df_sorted.columns else 0
                    st.markdown(f"""
                    <div class='graph-comment'>
                    <b>Efficiency Interpretation:</b><br>
                    • <b>Bars/Line:</b> Plant treatment efficiency<br>
                    • <b>Red dashed line:</b> Target efficiency (75%)<br>
                    • <b>Average Efficiency:</b> {avg_eff:.1f}%
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown("</div>", unsafe_allow_html=True)
                
                # ===== WATER QUALITY GRAPH =====
                wq_records = get_water_quality_records(selected_branch, limit=100)
                if not wq_records.empty and len(wq_records) > 1:
                    wq_sorted = wq_records.sort_values('datetime')
                    
                    st.markdown("<div class='graph-container'>", unsafe_allow_html=True)
                    st.markdown("<div class='graph-title'>WATER QUALITY TREND</div>", unsafe_allow_html=True)
                    
                    fig_wq, axes = plt.subplots(3, 1, figsize=(12, 10))
                    fig_wq.patch.set_facecolor('#87CEEB')
                    
                    wq_dates = wq_sorted['datetime'].dt.strftime("%m-%d %H:%M")
                    
                    axes[0].set_facecolor('#f0f8ff')
                    axes[0].plot(wq_dates, wq_sorted['raw_ntu'], label="Raw Turbidity", color="#e74c3c", marker='o', linewidth=2, markersize=4)
                    axes[0].set_title("Raw Water Turbidity", color="#1e3c72", fontsize=12, weight='bold')
                    axes[0].set_ylabel("NTU", fontsize=10)
                    axes[0].grid(True, linestyle='--', alpha=0.3)
                    axes[0].legend()
                    axes[0].tick_params(axis='x', rotation=45)
                    
                    axes[1].set_facecolor('#f0f8ff')
                    axes[1].plot(wq_dates, wq_sorted['outlet_ntu'], label="Treated Turbidity", color="#27ae60", marker='s', linewidth=2, markersize=4)
                    axes[1].axhline(y=5, color='#f39c12', linestyle='--', linewidth=2, label="Target 5 NTU")
                    axes[1].set_title("Treated Water Turbidity", color="#1e3c72", fontsize=12, weight='bold')
                    axes[1].set_ylabel("NTU", fontsize=10)
                    axes[1].grid(True, linestyle='--', alpha=0.3)
                    axes[1].legend()
                    axes[1].tick_params(axis='x', rotation=45)
                    
                    axes[2].set_facecolor('#f0f8ff')
                    axes[2].plot(wq_dates, wq_sorted['pump_Lh'], label="Pump Rate", color="#1e3c72", marker='^', linewidth=2, markersize=4)
                    axes[2].set_title("Pump Setting (L/h)", color="#1e3c72", fontsize=12, weight='bold')
                    axes[2].set_xlabel("Time", fontsize=10)
                    axes[2].set_ylabel("L/h", fontsize=10)
                    axes[2].grid(True, linestyle='--', alpha=0.3)
                    axes[2].legend()
                    axes[2].tick_params(axis='x', rotation=45)
                    
                    plt.tight_layout()
                    st.pyplot(fig_wq)
                    
                    avg_raw = wq_records['raw_ntu'].mean()
                    avg_outlet = wq_records['outlet_ntu'].mean()
                    avg_pump = wq_records['pump_Lh'].mean() if 'pump_Lh' in wq_records.columns else 0
                    st.markdown(f"""
                    <div class='graph-comment'>
                    <b>Water Quality Statistics:</b><br>
                    • <b>Average Raw Turbidity:</b> {avg_raw:.1f} NTU<br>
                    • <b>Average Treated Turbidity:</b> {avg_outlet:.1f} NTU<br>
                    • <b>Average Pump Setting:</b> {avg_pump:.1f} L/h<br>
                    • <b>Target Treated Turbidity:</b> 5 NTU (yellow dashed line)
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown("</div>", unsafe_allow_html=True)
                
            else:
                st.info("📝 No data available. Start adding records!")
            
            if has_data and not filtered_data.empty and st.session_state.user_role == "Admin":
                st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
                st.markdown("### Export Data")
                
                col_export1, col_export2 = st.columns(2)
                with col_export1:
                    if st.button("📊 Export Full Excel Report", use_container_width=True, key="full_excel_export_btn"):
                        with st.spinner("Generating comprehensive Excel report..."):
                            excel_data = export_full_excel_report(
                                selected_branch, 
                                st.session_state.selected_year,
                                st.session_state.selected_month,
                                st.session_state.selected_day,
                                st.session_state.filter_level
                            )
                            
                            if excel_data:
                                st.download_button(
                                    label="📥 Download Excel Report",
                                    data=excel_data,
                                    file_name=f"Nzove_Complete_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                                st.success("✅ Excel report generated successfully!")
                            else:
                                st.error("❌ No data available to export")
                
                with col_export2:
                    if st.button("📊 Export Simple CSV", use_container_width=True, key="csv_export_btn"):
                        try:
                            csv_buffer = filtered_data.to_csv(index=False)
                            st.download_button(
                                label="📥 Download CSV",
                                data=csv_buffer,
                                file_name=f"Nzove_Data_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv",
                                use_container_width=True
                            )
                        except Exception as e:
                            st.error(f"Export error: {str(e)}")

    # ================= MONTHLY TEAM EVALUATION =================
    elif st.session_state.menu_option == "Monthly Team Evaluation" and st.session_state.user_role == "Admin":
        st.markdown("## Monthly Team Performance Evaluation")
        st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
        
        eval_year = 2026
        eval_month = datetime.now().month
        
        col_y, col_m = st.columns(2)
        with col_y:
            eval_year = st.number_input("Year", min_value=2026, max_value=2026, value=eval_year, key="eval_year")
        with col_m:
            eval_month = st.selectbox("Month", range(1, 13), index=eval_month-1, 
                                     format_func=lambda x: ["January","February","March","April","May","June","July","August","September","October","November","December"][x-1], 
                                     key="eval_month")
        
        st.markdown("### 🏆 Team Rankings (Based on Revenue)")
        rankings = calculate_team_rankings(selected_branch, eval_year, eval_month)
        
        if rankings:
            rank_df = pd.DataFrame(rankings)
            rank_df = rank_df[["rank", "team", "revenue", "efficiency", "revenue_water", "cost", "records", "total_cost", "total_profit"]]
            rank_df.columns = ["Rank", "Team", "Total Revenue (RWF)", "Efficiency %", "Revenue Water", "Cost/m³", "Records", "Total Cost", "Profit"]
            
            def color_rank(val):
                if val == 1:
                    return 'background-color: #FFD700; color: black; font-weight: bold;'
                elif val == 2:
                    return 'background-color: #C0C0C0; color: black; font-weight: bold;'
                elif val == 3:
                    return 'background-color: #CD7F32; color: white; font-weight: bold;'
                return ''
            
            rank_df["Total Revenue (RWF)"] = rank_df["Total Revenue (RWF)"].apply(lambda x: f"{x:,.0f}")
            rank_df["Total Cost"] = rank_df["Total Cost"].apply(lambda x: f"{x:,.0f}")
            rank_df["Profit"] = rank_df["Profit"].apply(lambda x: f"{x:,.0f}")
            
            st.dataframe(rank_df.style.applymap(color_rank, subset=['Rank']), use_container_width=True)
            
            best_team = rankings[0]
            st.success(f"🏆 **Best Team: Team {best_team['team']}** with Revenue {best_team['revenue']:,.0f} RWF")
        else:
            st.info("No data available for ranking")
        
        st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
        
        st.markdown("### Select Team")
        selected_team_eval = st.radio("Choose Team", TEAM_TYPES, horizontal=True, key="team_eval_select")
        team_info = get_team_info(selected_branch, selected_team_eval)
        
        try:
            perf = calculate_monthly_performance_from_records(selected_branch, selected_team_eval, eval_year, eval_month)
        except Exception as e:
            st.error(f"Error: {str(e)}")
            perf = None
        
        if perf and perf["record_count"] > 0:
            st.markdown(f"### Performance Report - Team {selected_team_eval}")
            st.markdown(f"**Month:** {['January','February','March','April','May','June','July','August','September','October','November','December'][eval_month-1]} {eval_year}")
            st.markdown(f"**Records:** {perf['record_count']} days")
            
            st.markdown("#### Team Info")
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                st.markdown(f"""
                <div class='team-card'>
                    <b>Team Leader:</b> {team_info.get('team_leader', 'Not Assigned') if team_info else 'Not Assigned'}<br>
                    <b>Team Type:</b> {selected_team_eval}
                </div>
                """, unsafe_allow_html=True)
            with col_t2:
                st.markdown(f"""
                <div class='team-card'>
                    <b>Contact:</b> {team_info.get('contact', 'N/A') if team_info else 'N/A'}
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("#### Metrics")
            mc1, mc2, mc3, mc4 = st.columns(4)
            with mc1:
                st.metric("Avg Efficiency", f"{perf['avg_efficiency']:.1f}%", 
                         delta=f"{perf['avg_efficiency'] - 75:.1f}%", 
                         delta_color="normal" if perf['avg_efficiency'] >= 75 else "inverse")
            with mc2:
                st.metric("Raw Water", format_number(perf['total_raw']))
            with mc3:
                st.metric("Total Treated", format_number(perf['total_output']))
            with mc4:
                st.metric("Revenue Water", format_number(perf['revenue_water']))
            
            mc5, mc6, mc7, mc8 = st.columns(4)
            with mc5:
                st.metric("Backwashing", format_number(perf['backwashing']))
            with mc6:
                st.metric("Non-Revenue %", f"{perf['non_revenue_pct']:.1f}%")
            with mc7:
                st.metric("Cost per m³", f"{perf['cost_per_m3']:.1f} RWF")
            with mc8:
                st.metric("Profit", format_number(perf['total_profit']))
            
            if perf['avg_efficiency'] >= 75:
                st.success(f"Excellent! Team {selected_team_eval} achieved {perf['avg_efficiency']:.1f}% efficiency")
            else:
                st.warning(f"Team {selected_team_eval} efficiency is {perf['avg_efficiency']:.1f}%, below 75% target")
            
            if st.session_state.user_role == "Admin":
                if st.button("Save Evaluation", use_container_width=True):
                    eval_data = {
                        "year": eval_year, "month": eval_month, "branch": selected_branch,
                        "team_type": selected_team_eval,
                        "team_leader": team_info.get("team_leader", "N/A") if team_info else "N/A",
                        "avg_efficiency": perf["avg_efficiency"],
                        "total_raw": perf["total_raw"],
                        "total_output": perf["total_output"],
                        "revenue_water": perf["revenue_water"],
                        "backwashing": perf["backwashing"],
                        "non_revenue_pct": perf["non_revenue_pct"],
                        "total_cost": perf["total_cost"],
                        "total_revenue": perf["total_revenue"],
                        "total_profit": perf["total_profit"],
                        "cost_per_m3": perf["cost_per_m3"],
                        "compliance_score": perf["compliance_score"],
                        "rank": 1,
                        "remarks": "Good" if perf["avg_efficiency"] >= 75 else "Needs improvement"
                    }
                    save_monthly_evaluation(eval_data)
                    st.success(f"Evaluation saved!")
        else:
            st.info(f"No data for Team {selected_team_eval}")

    # ================= RECORDS =================
    elif st.session_state.menu_option == "Records":
        st.markdown("### Historical Log Register - Nzove")
        all_records = get_records(selected_branch)
        if not all_records.empty:
            all_records["Date"] = pd.to_datetime(all_records["Date"])
            display_records = all_records.sort_values("Date", ascending=False)
            display_records.index = np.arange(1, len(display_records) + 1)
            st.dataframe(display_records, use_container_width=True)
            
            if st.session_state.user_role == "Admin":
                try:
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        display_records.to_excel(writer, sheet_name="Records", index=False)
                    output.seek(0)
                    st.download_button("Download Excel", data=output, file_name="Nzove_Records.xlsx", 
                                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                                     use_container_width=True)
                except:
                    csv_buffer = display_records.to_csv(index=True)
                    st.download_button("Download CSV", data=csv_buffer, file_name="Nzove_Records.csv", 
                                     mime="text/csv", use_container_width=True)
        else:
            st.info("📝 No records found. System is empty. Start adding new records!")

    # ================= WATER QUALITY DATA =================
    elif st.session_state.menu_option == "Water Quality Data":
        st.markdown("## Water Quality Monitoring Data")
        st.markdown("<div class='custom-divider'></div>", unsafe_allow_html=True)
        
        wq_records = get_water_quality_records(selected_branch, limit=100)
        
        if wq_records.empty:
            st.info("📝 No water quality records found.")
        else:
            wq_records = wq_records.sort_values('datetime')
            st.markdown(f"### Water Quality Records - Nzove")
            st.markdown(f"**Total:** {len(wq_records)} records")
            
            tab1, tab2, tab3 = st.tabs(["Turbidity Trend", "Summary", "Records"])
            
            with tab1:
                if len(wq_records) > 1:
                    fig, ax = plt.subplots(figsize=(12, 5))
                    ax.set_facecolor('#f0f8ff')
                    fig.patch.set_facecolor('#87CEEB')
                    
                    wq_sorted = wq_records.sort_values('datetime')
                    wq_dates = wq_sorted['datetime'].dt.strftime("%m-%d %H:%M")
                    ax.plot(wq_dates, wq_sorted['raw_ntu'], label="Raw NTU", color="#e74c3c", marker='o', linewidth=2, markersize=6)
                    ax.plot(wq_dates, wq_sorted['outlet_ntu'], label="Treated NTU", color="#27ae60", marker='s', linewidth=2, markersize=6)
                    ax.axhline(y=5, color='#f39c12', linestyle='--', alpha=0.7, linewidth=2, label="Target 5 NTU")
                    
                    ax.set_xlabel("Date & Time", fontsize=12, color='#1e3c72')
                    ax.set_ylabel("Turbidity (NTU)", fontsize=12, color='#1e3c72')
                    ax.set_title("Nzove - Turbidity Trend", fontsize=14, weight='bold', color='#1e3c72')
                    ax.legend(loc='upper right', fontsize=11)
                    ax.grid(True, linestyle='--', alpha=0.3)
                    plt.xticks(rotation=45)
                    plt.tight_layout()
                    st.pyplot(fig)
                    
                    avg_raw = wq_records['raw_ntu'].mean()
                    avg_outlet = wq_records['outlet_ntu'].mean()
                    removal = ((avg_raw - avg_outlet) / avg_raw * 100) if avg_raw > 0 else 0
                    st.markdown(f"""
                    <div class='graph-comment'>
                    <b>Turbidity Statistics:</b><br>
                    • <b>Red line:</b> Raw water turbidity<br>
                    • <b>Green line:</b> Treated water turbidity<br>
                    • <b>Yellow dashed line:</b> Target (5 NTU)<br>
                    • <b>Average Raw:</b> {avg_raw:.1f} NTU<br>
                    • <b>Average Treated:</b> {avg_outlet:.1f} NTU<br>
                    • <b>Removal Rate:</b> {removal:.1f}%
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.info("Need at least 2 records for trend")
            
            with tab2:
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Avg Raw Turbidity", f"{wq_records['raw_ntu'].mean():.1f} NTU")
                    st.metric("Avg Treated Turbidity", f"{wq_records['outlet_ntu'].mean():.1f} NTU")
                    st.metric("Avg Efficiency", f"{wq_records['efficiency'].mean():.1f}%")
                with col2:
                    st.metric("Avg Dosage", f"{wq_records['dose_mgL'].mean():.1f} mg/L")
                    st.metric("Avg Pump Rate", f"{wq_records['pump_Lh'].mean():.1f} L/h")
                    st.metric("Avg pH", f"{wq_records['raw_ph'].mean():.1f}")
            
            with tab3:
                display_wq = wq_records[['datetime', 'flow_m3', 'raw_ntu', 'outlet_ntu', 'efficiency', 'pump_Lh', 'raw_ph', 'raw_temperature']]
                display_wq.columns = ['Date & Time', 'Flow (m³/h)', 'Raw NTU', 'Outlet NTU', 'Efficiency (%)', 'Pump (L/h)', 'pH', 'Temp (°C)']
                st.dataframe(display_wq, use_container_width=True)
                
                if st.button("Export Water Quality", use_container_width=True):
                    csv_buffer = wq_records.to_csv(index=False)
                    st.download_button("Download CSV", data=csv_buffer, file_name="Water_Quality_Nzove.csv", mime="text/csv")